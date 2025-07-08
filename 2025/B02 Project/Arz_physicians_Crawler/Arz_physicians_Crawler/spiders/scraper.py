import os, glob,csv
import random
from datetime import datetime
from collections import OrderedDict
from playwright.sync_api import sync_playwright

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

# Global set to track unique record keys
seen_record_keys = set()

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
input_folder = os.path.join(project_root, 'input')
logs_folder = os.path.join(project_root, 'logs')
output_folder = os.path.join(project_root, 'output')

os.makedirs(logs_folder, exist_ok=True)
os.makedirs(output_folder, exist_ok=True)

# Timestamp and logs
timestamp = datetime.now().strftime("%d%m%Y%H%M")
log_file_path = os.path.join(logs_folder, f'Arztsuche_logs_{timestamp}.txt')
seen_postcodes_path = os.path.join(output_folder, 'seen_postcodes.csv')

# Define the single output file path (timestamped once for the run)
output_file = os.path.join(output_folder, f"All_Postcode_Records_{timestamp}.csv")

# API Headers
api_headers = {
                    'Accept': 'application/json, text/plain, */*',
                    'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
                    'Authorization': 'Basic YmRwczpma3I0OTNtdmdfZg==',
                    'Cache-Control': 'no-cache',
                    'Connection': 'keep-alive',
                    'Pragma': 'no-cache',
                    'Referer': 'https://arztsuche.116117.de/',
                  }

def write_log(message):
    with open(log_file_path, mode='a', encoding='utf-8') as file:
        file.write(f'{message}\n')
    print(message)

def write_csv(all_records, code):
    if not all_records:
        return

    headers = all_records[0].keys()  # Get headers from the first record
    with open(output_file, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)

        # Check if file is empty to write headers
        if csvfile.tell() == 0:
            writer.writeheader()

        # Write all records
        writer.writerows(all_records)

    write_log(f"[INFO] Saved {len(all_records)} records for postcode {code}")

def load_seen_postcodes():
    if not os.path.exists(seen_postcodes_path):
        return set()
    with open(seen_postcodes_path, mode='r', encoding='utf-8') as f:
        return set(row['Plz'] for row in csv.DictReader(f))

def mark_postcode_seen(postcode):
    # Read existing postcodes if file exists
    existing_postcodes = set()
    if os.path.exists(seen_postcodes_path):
        with open(seen_postcodes_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            existing_postcodes = {row['Plz'] for row in reader}

    # Skip if postcode already exists
    if postcode in existing_postcodes:
        return

    # Write new postcode
    file_exists = os.path.isfile(seen_postcodes_path)
    with open(seen_postcodes_path, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['Plz'])
        if not file_exists:
            writer.writeheader()
        writer.writerow({'Plz': postcode})

def get_clean_record(data):
    item = OrderedDict()
    try:
        item['Full Name'] = data.get('name', '').strip()
        item['Title'] = data.get('anrede', '').strip()
        item['Specialty/Fachgebiet(e)'] = ', '.join([ag.get('value') for ag in data.get('ag', [])])
        item['Address'] = ' '.join(filter(None, [data.get('strasse', '').strip(), data.get('hausnummer', '').strip()]))
        item['ZIP Code'] = data.get('plz', '').strip()
        item['City'] = data.get('ort', '').strip()
        phones = [data.get(key) for key in ['tel', 'handy', 'fax'] if data.get(key)]
        item['Phone Number(s)'] = ', '.join(phones)
        item['Email Address'] = data.get('email', '').strip()
        item['Website'] = data.get('web', '').strip()
        item['Languages spoken'] = ', '.join(data.get('zm', [])) or ', '.join(data.get('fs', []))
        item['Type of practice'] = 'Arzt' if data.get('arzt') else 'Praxis'
        kvg_values = [v for kvg in data.get('kvg', []) for v in kvg.get('values', [])]
        item['Affiliated health insurance'] = ', '.join(kvg_values)
        item['Chamber/registration information'] = data.get('kv', '')

        day_translation = {
            'Mo.': 'Mon.',
            'Di.': 'Tue.',
            'Mi.': 'Wed.',
            'Do.': 'Thu.',
            'Fr.': 'Fri.',
            'Sa.': 'Sat.',
            'So.': 'Sun.'
        }

        # Transform data
        office_hours = []
        for entry in data.get('tsz', []):
            german_day = entry.get('t', '')
            english_day = day_translation.get(german_day, german_day)
            office_hours.append({'day': english_day, 'time': entry.get('d', '')})

        item['Office hours'] = ' || '.join([f"{entry['day']} {entry['time']}" for entry in office_hours])

        # Create a unique key for duplication check
        key_parts = [item['Full Name'], item['ZIP Code'], item['Phone Number(s)']]
        record_key = '|'.join(key_parts).lower()

        return item, record_key

    except Exception as e:
        write_log(f"[ERROR] Error while parsing record: {e}")
    return item

def accept_cookies(page):
    try:
        if page.locator('div#dialog1').is_visible(timeout=5000):
            page.click('button[name="cookieConsentAcceptAll"]')
            write_log("[INFO] Accepted cookies.")
    except Exception as e:
        write_log(f"[INFO] Cookie dialog not shown or already accepted: {e}")

def fetch_postcode_data(code, proxies):
    try:
        proxy_ip = random.choice(proxies)
        with sync_playwright() as p:
            # browser = p.chromium.launch(headless=False)
            browser = p.chromium.launch(
                headless=True,
                proxy={"server": f"http://{proxy_ip}"}
            )
            context = browser.new_context()
            page = context.new_page()
            page.set_default_timeout(60000)

            page.goto("https://arztsuche.116117.de/")
            page.wait_for_load_state('networkidle')
            accept_cookies(page)
            page.fill('#SuchortSearchInput', code)
            page.wait_for_selector('.option-item', timeout=5000)
            page.click('.option-item')
            page.click('#UmkreisRadioSelect__BV_toggle_')
            page.wait_for_selector('#select-Umkreis-item-check-4')
            page.click('#select-Umkreis-item-check-4')

            def is_valid_response(response):
                return "https://arztsuche.116117.de/api/data" in response.url and response.status == 200

            with page.expect_response(is_valid_response) as response_info:
                page.click('#searchBtn')

            response = response_info.value
            mark_postcode_seen(code)
            return response.json() if response else {}
    except Exception as e:
        write_log(f"[ERROR] Failed fetching browser data for {code}: {e}")
        return {}

def search_postcode(code, proxies):
    if not code:
        return

    try:
        all_records = []
        response_data = fetch_postcode_data(code, proxies)

        for record in response_data.get('arztPraxisDatas', []):
            cleaned, record_key = get_clean_record(record)
            if cleaned:
                if record_key in seen_record_keys:
                    print('already scraped so skipped', record_key)
                    continue
                seen_record_keys.add(record_key)
                all_records.append(cleaned)

        if all_records:
            write_csv(all_records, code)
            mark_postcode_seen(code)
        else:
            mark_postcode_seen(code)
            write_log(f"[INFO] No records found for postcode {code}")

    except Exception as e:
        write_log(f"[ERROR] Exception during processing postcode {code}: {e}")

def read_input_postcodes():
    print('Reading postcode files from input directory...')
    postcodes = []

    # Read from PLZ.xlsx
    try:
        for xlsx in glob.glob(os.path.join(input_folder, 'PLZ.xlsx')):
            sheet = load_workbook(xlsx, read_only=True).active
            for row in sheet.iter_rows(min_row=2, max_col=1):
                cell = row[0]
                if cell.value:
                    val = str(cell.value).strip()
                    if cell.number_format == '00000' and val.isdigit():
                        val = f"{int(val):05d}"
                    if val not in postcodes:
                        postcodes.append(val)
    except Exception as e:
        write_log(f"[ERROR] Reading Excel file: {e}")

    try:
        filepath = glob.glob(os.path.join(input_folder, 'german-postcodes.csv'))[0]
        with open(filepath, mode='r', encoding='utf-8') as f:
            records = [row.get('Plz', '') for row in csv.DictReader(f, delimiter=';')]
            postcodes.extend(records)
            a=1
    except Exception as e:
        write_log(f"[ERROR] Unable to read input postcode file: {e}")

    if not postcodes:
        write_log("No postcodes found in input files.")

    return list(set([p for p in postcodes if p.strip()]))

def read_input_from_file():
    file_path = glob.glob(os.path.join(input_folder, 'proxies.txt'))[0]
    try:
        with open(file_path, mode='r') as txt_file:
            return [line.strip() for line in txt_file.readlines() if line.strip()]

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return []


if __name__ == '__main__':
    write_log("[INIT] Script started")
    all_postcodes = read_input_postcodes()
    seen_postcodes = load_seen_postcodes()
    proxies = read_input_from_file()

    for postcode in all_postcodes:
        if len(str(postcode)) < 5:
            write_log(f"[SKIPPED] Invalid postcode '{postcode}': Must be 5 digits. Skipping entry.")
            continue

        if postcode in seen_postcodes:
            write_log(f"[SKIP] Postcode {postcode} already processed.")
            continue

        search_postcode(str(postcode), proxies)

    write_log("[COMPLETE] Script finished")