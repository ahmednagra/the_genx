import collections
import glob, json, time, logging, html, re
import string
from math import ceil
from datetime import datetime
from collections import OrderedDict
from uuid import uuid4

import requests
from scrapy import Spider, Request, signals

from scrapy.exceptions import CloseSpider
from openpyxl.reader.excel import load_workbook

logger = logging.getLogger("ehealth_scraper")
logger.setLevel(logging.INFO)

class EhiSpiderSpider(Spider):
    name = "ehi"
    allowed_domains = ["www.ehealthinsurance.com", "graph.ehealthinsurance.com"]

    custom_settings = {
        # 'COOKIES_ENABLED': True,
        'OFFSITE_ENABLED': False,
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 408],
        'CONCURRENT_REQUESTS': 1,
        'FEED_EXPORTERS': {
            'xlsx': 'scrapy_xlsx.XlsxItemExporter',
        },
        'FEEDS': {
            f'outputs/Ehealth Records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx': {
                'format': 'xlsx',
                'fields': ['Zip Code', 'County', 'Effective Date', 'State', 'Competitor Name', 'Plan Name',
                           'Premium Per Employee','Percentage for dependents 0%', 'Plan type', 'Annual Deductible (per person)',
                           'Out-of-Pocket limit',
                           'Health Savings Account (HSA) eligible', 'Primary Doctor', 'Specialist',
                           'Generic/Tier 1 Drugs',
                           'Preferred Brand/Tier 2 Drugs', 'Non-Preferred Brand/Tier 3 Drugs:',
                           'Specialty/Tier 4 Drugs',
                           'Emergency room', 'Urgent care', 'X-rays', 'CT,PET scans, MRIs',
                           'Mental Health Outpatient Office Visit', 'Psychiatric hospital stay', 'URL']
            }
        },

        'ZYTE_API_EXPERIMENTAL_COOKIES_ENABLED': True,

        'DOWNLOAD_HANDLERS': {
            'http': 'scrapy_zyte_api.ScrapyZyteAPIDownloadHandler',
            'https': 'scrapy_zyte_api.ScrapyZyteAPIDownloadHandler',
        },

        'DOWNLOADER_MIDDLEWARES': {
            'scrapy_zyte_api.ScrapyZyteAPIDownloaderMiddleware': 1000,
            'scrapy_poet.InjectionMiddleware': 543,
        },

        'REQUEST_FINGERPRINTER_CLASS': 'scrapy_zyte_api.ScrapyZyteAPIRequestFingerprinter',
        'TWISTED_REACTOR': 'twisted.internet.asyncioreactor.AsyncioSelectorReactor',
        'ZYTE_API_TRANSPARENT_MODE': True,
        'ZYTE_API_KEY': 'bd92c1e24ac246a58231c2b7741a1120',
    }

    cookies = {
        'ehiVisitorId': 'C4667D4488814783A61F1BCFDA83C99C.ifp-c3po-app-prod-rs-v132-xbhcg',
        'locale': 'en',
        'ehi_newhouse': 'true',
        'visitedProductLine': 'SMB',
        'groupId': '7543132',
        'groupUUId': '38173DFBFFC1C08DE0631814E70A1179',
        'ehi_cc-id': '49a39bb4-58fd-42ad-a19d-16a1eec616ab',
        'showChatBar': 'false',
        'zipCode': '54301',
        'ehi.alliance_id': 'ehe12862000',
        'rc_aid': 'ehe12862000',
        'bm_ss': 'ab8e18ef4e',
        'ak_bmsc': 'F03FB2EA7F900AE0F383A09F3DBCF13A~000000000000000000000000000000~YAAQi89YaCydCcGXAQAAbgJF2xzkVsNLqPs+CMkqRwt41s/L/r3z4KceQ1WAc6W7xy0SfA0VySIAONa6qVxbES4YeaxwoxPIkYYODuSgPK9WV8LtQlg9CJSQ5qg8JE2gNnUp84dddncI6UTS1mIoRhNKgwN2nOVIvSuaRuWLTyo+lnCdlpu2G7WFV/Xm18/cAU1pqgAD3LRIxdPxrGEybM/TFpXFF5ZGIgyFJbroQ1sHmQY1imPpTD9ZTM7hPb/GDdVpOAZ7Ebyb8yUg45B2LIwlUt91Z3jJpB8UGafaer0kJr4Szo2jsxO1kvpYCBdrNjWzEuT8TtHRZqZYaRA4nhWnHKZP/3eLAA5L/YGgKA6UwZ1mGVtQGwm0j5rIPz/M2uzvtuGJEOAKT0dYJd/nMecAjyI=',
        'at_check': 'true',
        'AMCVS_A821776A5245B31A0A490D44%40AdobeOrg': '1',
        'ehi-app-route': 'c940a9ec0daab6b41c31217f8cd80104|ca560dc112de7419cbba1722de74949d',
        'JSESSIONID': 'BFA046FEFD88E80DC5D4857277608AE6',
        'ehiCookie': '"U0hPV19CUk9XU0VSX1dBUk5JTkc9TnxfRW50cnlVUkw9L3xfV2ViQ29udGV4dD1EaXJlY3Q="',
        'AKA_A2': 'A',
        'OptanonConsent': 'isGpcEnabled=1&datestamp=Sat+Jul+05+2025+20%3A54%3A42+GMT%2B0500+(Pakistan+Standard+Time)&version=202503.2.0&browserGpcFlag=1&isIABGlobal=false&consentId=49a39bb4-58fd-42ad-a19d-16a1eec616ab&identifierType=Cookie+Unique+Id&isAnonUser=0&hosts=&interactionCount=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0003%3A1%2CC0004%3A1%2CC0002%3A1%2CSSPD_BG%3A0%2CC0005%3A0&AwaitingReconsent=false',
        'mbox': 'PC#1cfc970d01284bae8c98bb0afabe49ee.35_0#1813941601|session#34cb4cbbe5324c708dd6b000320b901a#1751734350',
        'bm_lso': 'BD80972D63DEA6C65C05BA3B53C2C1EC9B922DEEC74D6D7E915BDC9215661822~YAAQi89YaKqaFcGXAQAA/ntk2wRVZKTHs/DtnoSIWnhvBE22x6+Xtj0WmpIb5cK2ThIj+mrz/aWKnQS8bmyNktSXa6gPHn+sEHZ4A8Ylrrblg49cdBeaz/f6RrGmQF50RxCig3rnTvh11K0urN8Uky2l5LuWIImLkMw0VyNbv58HW402Nf+/iSQzXZH+l93Ok51e1qvF6u51ff1GzXo+A7mq0k7AvL3cPHA/wnVVX1OmBu942yngFfX/rxZhA+OQFWS70KrT86KfXHORCg2SQ89KuPyDN6cckzPrqTowLfgsped89YMW7DklXhNFurJSNY9mxvwFRwOWK4/NIcqteOP7h9w0/jypnB6eXXF9h5ZDMiEN0AsGy0JRwmlOJaLUu8NByC9KjPGSV7aAUSjjX+JaRZ4nHeUlqKYBsgofxwuXRlewPbt2bVgJRTwsFms7/Fd2tYdzseyhm7sGKc6VXUGAHVT3mOvkxQ==^1751732496453',
        'bm_s': 'YAAQi89YaL8kFsGXAQAAU5hl2wO+TktdIH9PqTyjHOSduCNTm/MV/GQ7Toh8gy1Zqmb5bzeTHbLhc7T0R7uhltGIbGufDCTrWxwIEHmtXlrE6UGhixRiQr3MJtjOC6/BVYlJ32VcCy+hJK+p/NLfnuDB5czQFV25C9Hyc3u/JYWC1KzfnB85csxlaXp8qTMJdG4GF23o/e3iTEYsWxcCUN240bY3g0CBbtLvcJv8U0Mq+8yHA9SRLTXmxl7+NqQxwbFxI1vjIM9EvWJvKOhNl6jT69CYZtTo2H2g4/WwZfXsxlADFVzm1fep9fheufFqr2L6+c15UYdsdyRlxPF6ElX+zsZUMU3jQyEDqHypLSNAp75x6RCYxVNaHgl5ZcJWUhnuctQqiYxRdCJf1S6C0SU0DBjAjkAHu7CZ6WN6TROaxmJ0bHHYJj/F5WCLCH84phmkywHeh8ugDOwymr0QVfsYMz6/AJXW6Zd3OGJW4bAGLejaE+yfM+8XJzW5FGN2wEVX7hkwsUqSESm1oGzwU1RyIxmWxNSmhns+KRvdPgUpd/Qf0Wmv3FHusHAxpX1ud9b4XGZu2zXhJ9kebHclZYOJclxpdaQ=',
        'bm_so': 'B4827C98AD6419B5D65AC40CE0F437B95B22333DAB54C30E5DF36C30A4BDCFF9~YAAQi89YaMAkFsGXAQAAU5hl2wSZkjdoRP/5hzDHd2Yq1KyXOtqUN5AtNL9Yt8cILiEkVVT0qxb3PEUdXhr072JPXz9wvzwqvCdJd7QB0C3IfNHa5MusCKx7MiF3dfRp350YyqGkwIFAVtZPQcoCeJjkz4W/CXXMcaJvP8gV7bldsiTmyMocI5PmPXy18RS5prRVIiykFYthn92U5NF1zCmxycyg+8r9ez8S+7FceCGSB4Ocyx4x1XPXwAFu83Wh2zfaiJE22K7685xGMJkn0koepEIFnexuPe47WJ8q1FJPh2rNKFQCaMcPEYZx6av/m8L50BVfph8RWMnK0gjxBkFCa3j9D6L9oMbX8KbqtNPjExL9PWdf+06uoA/jCfk9zMPEfmydJmoAiScUk+t28BuseNLUhHDw6mXOkNjEHIlc5Cly/4J5fzezPejJwdkI7SfJQ39RKe1U00ypYQiWiLrSPFjq8tLJSA==',
        'bm_sz': '050CCD0C1FA574FEE5BA03A67C8C1CF5~YAAQi89YaMIkFsGXAQAAU5hl2xxwtIE52/z2Er74NymIQShkchBUMOp/myyLw/fYUOFJUjwkmF7C2WR6hF0SmcT0SoecEMqURn5lixqKoY28jbJj//cF3qCIR4uQ/C2U8CIKAJdLzaNDwu4JrJSv0dPcTfB72QjOoaXdb8HwwqQq104vdPXMLifILAkLkHNbjz7Xv0+UTOIpPm+DF488vvtqEQJ90a4uq4uZIqQeKK3eQ5aQIPdopF4Ov6dW6zip8/ugY9MpEkwLNwulBGPRIeQ7oOPCiCBGHrgj2YyyzME8xeiNGesQLbCteV+GYCi+BpG4/kejGcnoygIv/3/KXtF4soGnr8YJetehEPWodmxHt3aDfMcNMW3OqE09cUC0NKmqH4EfCJxHbk8JGA8bMkmBDjEyUKVWlsMoNiF5zQ7gLMeZFLm3pQiXy6BBeur8A3F5O2HfWJ4Q9j+EdeOG2115/v1FCrEDtM0s8F1pPA==~3553077~3688006',
        '_abck': 'E0800626ED37547232A8E05277B62C8F~0~YAAQi89YaEgmFsGXAQAA35tl2w5qVIyOVPholkT2DxECMQFRvjltLonsULfWET8ivfskwkzXB/R5dIQW+KkTs6ZZN6z2b55w8hrsJ0rw7CsKuKT5zt8H8+1iuQhS5ejYiHwCMq8ecY5Mrs2pw1t48J5Da7AG2Ne3sctgeDHv7eWgWRqX27mF2C2Leqm+cK3PjQhDHMBDMcHegMgVLvhv4r0WP5+B1n+VPu5uN1JDbJjGTkzermpj3Lscp5Sgag4llV5pSCj3zgcOjdzBqnX96kcMrqr0ZcZwKOf19LWB125FYDpogZs2K6zB0rr7+1sEKQjEm2u6nCDuNgjLJFQE9ui3ie7+3RZNkKabLs0ZpUKOUz91u+3xD5DmplIccYb6+hO+C/NK+FAncwLie6TgZ4OeQRwpqsFkZ+Zru+6Gjwh7ksMkC0xkVl2o43e4MYZy3nLytebAKsAZCmwkeDJwU1Fhtz7zDNuebOju9XJwm6aH0P6wAKCkxy9uZWeV2z7WLQfg2BTgpyQnC3lznaAO2gQIPHARK6Djq/vikNdM8mWDzoVkNUj3/gBWvkVmHwWCvrZXRP5DuGsYg3u/5LZmvQuG16WE7/e6QKK4Hst5ZQnZaoWH+bBOfPa2VkCT1bEWL3IruY20lf9rZMAjbmpUVK5lnGAfYkuHvKTlOAGKWJoFggYh/aFosw0ARyDwil7ZW7TgLiUtd4Te4JQ=~-1~-1~-1',
        'akavpau_wr': '1751732593~id=c801fbfcfbc4d429389c625683e29ed3',
        'AMCV_A821776A5245B31A0A490D44%40AdobeOrg': '179643557%7CMCIDTS%7C20275%7CMCMID%7C50986452785004399619220660008648306542%7CMCAID%7CNONE%7CMCOPTOUT-1751739764s%7CNONE%7CvVersion%7C5.5.0',
        'bm_sv': '9927F5F4F5739A37390808C2E946A3F3~YAAQi89YaJsyFsGXAQAAErRl2xzjzt2t7lL9Rkq91JYAJr0WE6Lhrgq70EIRnqjfWo4s4inMaBjhC9POYmuMp0DeK0du5HoYHlwRiUhjGB0uMZjVQAA7Kv3O02fYSfEApe9vZ5BRtLrU2OWSJCXn5v3vzqGYc1ejBD4wXRZQn/rcPAXt/lRBVhY2p7H4P70XbfBUhWeJzn5VS8nEmnyKtcm6A4Df5hUgvkqnUuCdi+kZa/eORP3BAegnBnSiNrecQjEcQKRZwOWkakM=~1',
    }

    census_cookies = {
    'ehiVisitorId': 'C4667D4488814783A61F1BCFDA83C99C.ifp-c3po-app-prod-rs-v132-xbhcg',
    'locale': 'en',
    'ehi_newhouse': 'true',
    'visitedProductLine': 'SMB',
    'groupId': '7543132',
    'groupUUId': '38173DFBFFC1C08DE0631814E70A1179',
    'ehi_cc-id': '49a39bb4-58fd-42ad-a19d-16a1eec616ab',
    'showChatBar': 'false',
    'zipCode': '54301',
    'ehi.alliance_id': 'ehe12862000',
    'rc_aid': 'ehe12862000',
    'ak_bmsc': 'F03FB2EA7F900AE0F383A09F3DBCF13A~000000000000000000000000000000~YAAQi89YaCydCcGXAQAAbgJF2xzkVsNLqPs+CMkqRwt41s/L/r3z4KceQ1WAc6W7xy0SfA0VySIAONa6qVxbES4YeaxwoxPIkYYODuSgPK9WV8LtQlg9CJSQ5qg8JE2gNnUp84dddncI6UTS1mIoRhNKgwN2nOVIvSuaRuWLTyo+lnCdlpu2G7WFV/Xm18/cAU1pqgAD3LRIxdPxrGEybM/TFpXFF5ZGIgyFJbroQ1sHmQY1imPpTD9ZTM7hPb/GDdVpOAZ7Ebyb8yUg45B2LIwlUt91Z3jJpB8UGafaer0kJr4Szo2jsxO1kvpYCBdrNjWzEuT8TtHRZqZYaRA4nhWnHKZP/3eLAA5L/YGgKA6UwZ1mGVtQGwm0j5rIPz/M2uzvtuGJEOAKT0dYJd/nMecAjyI=',
    'at_check': 'true',
    'AMCVS_A821776A5245B31A0A490D44%40AdobeOrg': '1',
    'ehi-app-route': 'c940a9ec0daab6b41c31217f8cd80104|ca560dc112de7419cbba1722de74949d',
    'ehiCookie': '"U0hPV19CUk9XU0VSX1dBUk5JTkc9TnxfRW50cnlVUkw9L3xfV2ViQ29udGV4dD1EaXJlY3Q="',
    'bm_ss': 'ab8e18ef4e',
    'mbox': 'PC#1cfc970d01284bae8c98bb0afabe49ee.35_0#1813941601|session#f43cd31d02d64fb3a113c8296b526581#1751736480',
    'JSESSIONID': 'F2DF6A2841348F489DD8324FE811B4C7',
    '_abck': 'E0800626ED37547232A8E05277B62C8F~0~YAAQi89YaP3/JMGXAQAAqAeT2w5AE59P2wPKJe4XN6juKxJGlP1r9sI4Xvda3rM6L8bC73h2H+88BIu87hCUCKaoAVBXNrD/OXYFnEiMVg98jnAfe7Y0IqUg8ojc8ugel7hvyUC2lZiXaZInchl33p06s3ep4VT8wX9r5FHiiVyM9IQo55zo5GE2/GF9DGVrB4PIbnM7JYvaC6OfnkzZWbgy7FZL8OfzF3Q+ZGfpV4TZVfYeA6Ni8isK2xDHqNQAE0ZMcPO7Ab4pUWdRJ8HLJh2qDj4lmi59adveKuN51N4gHPZZ22WJKZsnhDnPpbWzFQbhzv1dD9RL1I6Bao2uMUYwVkMXU56jFn7hFptel5+36eFDGB3e+8wQG5XhYtmz5otf24GFwG8TgvftUF0EoHDwJi9tHgCpefS53iRqizTHN+kN6y4ekSkOkidQ1o6FtLF90rxu3O+hcjnRdugLlETycxINB8vXJMERmPxM+8IDjhAlJQBb/vh1jdkMdkuNFmeuKWFiK5TEEqp5vTJ0yn2R/odtry0fma9Bduvngv2U7WPi/r+0/uQiM9fggyJQ0gMZalupV58TRzpdfprM/SzdUk+7pRGGlk3gXQwHJZQSpAKWVPN4niUQjiSCD1WQ6EeIHNTfJ7qiFfRG/13t5rxkYKmfKFoZ3kJPZ9JxUJLYcBADZ2gFTVZ/+e/LKeh2x6MwtfRxmTa3ORKmfjuf4OgXSAhVpONVAJ/ryLtHNW/4jFv0iQ==~-1~-1~-1',
    'OptanonConsent': 'isGpcEnabled=1&datestamp=Sat+Jul+05+2025+22%3A13%3A35+GMT%2B0500+(Pakistan+Standard+Time)&version=202503.2.0&browserGpcFlag=1&isIABGlobal=false&consentId=49a39bb4-58fd-42ad-a19d-16a1eec616ab&identifierType=Cookie+Unique+Id&isAnonUser=0&hosts=&interactionCount=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0003%3A1%2CC0004%3A1%2CC0002%3A1%2CSSPD_BG%3A0%2CC0005%3A0&AwaitingReconsent=false',
    'bm_s': 'YAAQi89YaBulJcGXAQAAGuKU2wMemff8rryAwzm7HA5OtZQdHMtO8PG5GMIaAESHlBS5fdQ50V43ENhfSGA1cuOYkvVPn7vja/N7liHBP9lP4DhIWEYwLsSdWMaIcmlBkErSzNInkigLcuZT1R/MUvCLSnhBiYzAeNd+wXjW0a6T37LAGko6aGnrmCi+ALVOliMDR11a4py+DNKq5jVWjuWVYWngw0+2c8ZQPPtoYuNzR6iecN8AqvsGBNMdrK53DC8kh/BQtQ9XrH4hMJBF1khTaMq4owZ1d909qLPbBb0fCy00hawuaVx62TchWM0xaCtNjBVSssl9+Ie+XJ1O2vIlkwLrISYgrsm/lmma6EIqjXhnjTuF0E1Y+XYSDY0If31bW7dLVbJs/yBBAZLcwfpaglX8RNWoRbKjhAQsgWwanzket5BL+0scJ+LrbpekMVdpRUBz7UFy9MXlvaPvFHi0fLu5DidZCh9zesnBTHV1ecxnC878c66u4qHeTFy923GOS82XHMwY2eFc2XfcvTD9DuRF3ml33IJZBWBP4hq45i2B19UNnVYDBUQ4QqRrU+QwSmfRHtPJZQVyUxi7O62UihQZkFM=',
    'bm_so': 'A6EB79649BD037A032527495FB616AC98C47B56F3A8D5BC7DBD21A2D3C228798~YAAQi89YaBylJcGXAQAAGuKU2wTZsrSR019DDpR11hSWAngJFqONz7xE4crzjTLS42n9GmozATJxTY9wUgYxYLsmEbMU/teFiLcC6oQb4E8rDHBeM4B1/9tczglM5gNvWysZeL/J88D7R1YvbN2XPIA+YKraLCixqO5ADHNTU/NGrrmubrtHLhRZic13k65oepukTAFsn7tRiIRLO3jQUPLO6HRIluhiRrIQpTYCKW5R3De+ZA+QbhQg9gXJ1ZH3MfzWlLhnlIx4Q9uRrfzmBoW/Q6uS52RKL2gwUKfiXvUNfjjKQGYChXDrNBKdcc2qfRigCrEyo8pGsxsfWTK4B4nGyWkSYRLYw2nSVfz4JRevIL6a6AMVxI5qpAo3bQwII0KixjpJDmDCbqfXRL1rtfzfdD3/Z91+bjBhUpUpxSplwlQllMzLfQ+u1VCMMTqvDD32eI+EuU5RdiZZpzmm3Y04YgZb+R0gYw==',
    'bm_sz': '050CCD0C1FA574FEE5BA03A67C8C1CF5~YAAQi89YaB6lJcGXAQAAGuKU2xw9EsWHJBbJ/emFp8J5FmYEaF3QnJCw+d6SLNXAfGBtcs6Aog4bl+CLXUcE9DjRij/W+7DhcjMstAUe9gTkjn/owKDKFkY4viP8efOiZg50ho/vbBw/s0rrQJrytLLgmkWHUmz0aOQGzSVqXcPGinxWrvU4AdHCEabT8dvRXaEKitJ0JJrVGayQp2py18CkDnt+5JkFQWvlm8mcGsi0knZBtD3HAlHjtxNPcs/RIoFuBm0gvxYtq9xZsHfwUGsdtxldC7XPv9DnclmcYbGh20WdKRgPEOOb3FL7x5qQOoreIy1lGnJ1/VuXnm8KLJKxCs2Gwt6x+2TCbEyBrVgRauuIPbSVw5KtNMxpiGU5Owow9p5VLr8KOn4iDcOCKR2eTn/tNm94XXBQ5X1LyFoCXXDGrHb3sIeapva99r561TQu94mdeieLjIDppNw4KyBzQDOB0Q49K8n+1vLy+6jpN+WNaas4DwV1veJNfYQdb8uJ5EBcx/+ICbneksy/DClyZ7Yph1RHWP304bS1+/sS3ZNNHyYG~3553077~3688006',
    'AMCV_A821776A5245B31A0A490D44%40AdobeOrg': '179643557%7CMCIDTS%7C20275%7CMCMID%7C50986452785004399619220660008648306542%7CMCAID%7CNONE%7CMCOPTOUT-1751742858s%7CNONE%7CvVersion%7C5.5.0',
    'bm_lso': 'A6EB79649BD037A032527495FB616AC98C47B56F3A8D5BC7DBD21A2D3C228798~YAAQi89YaBylJcGXAQAAGuKU2wTZsrSR019DDpR11hSWAngJFqONz7xE4crzjTLS42n9GmozATJxTY9wUgYxYLsmEbMU/teFiLcC6oQb4E8rDHBeM4B1/9tczglM5gNvWysZeL/J88D7R1YvbN2XPIA+YKraLCixqO5ADHNTU/NGrrmubrtHLhRZic13k65oepukTAFsn7tRiIRLO3jQUPLO6HRIluhiRrIQpTYCKW5R3De+ZA+QbhQg9gXJ1ZH3MfzWlLhnlIx4Q9uRrfzmBoW/Q6uS52RKL2gwUKfiXvUNfjjKQGYChXDrNBKdcc2qfRigCrEyo8pGsxsfWTK4B4nGyWkSYRLYw2nSVfz4JRevIL6a6AMVxI5qpAo3bQwII0KixjpJDmDCbqfXRL1rtfzfdD3/Z91+bjBhUpUpxSplwlQllMzLfQ+u1VCMMTqvDD32eI+EuU5RdiZZpzmm3Y04YgZb+R0gYw==^1751735659509',
    'akavpau_wr': '1751735927~id=7815073a291b8abde994e00d019254a6',
    'bm_sv': '9927F5F4F5739A37390808C2E946A3F3~YAAQi89YaMrtJsGXAQAAcImY2xyepq3jOVivDcsO9IS/OxG2yYeu9Taw2XLFt6HdHk+KGnE+k/ZKIK/F56bXXRn7/15vo8+EZPJbMV7PSXcBfBKDNVYvdLSSJzkpB/GhxA/DuTo3rTmjJJkPt4tlAAyaiRTSsXAEs8KiIj0QIGem+wjGpHtK/dUnUztkfWnU9G0bPaiAjnd7dG4nNaz7H5QWn1siRkLW/vhgVT624OtT4/A8Zq6sM8kQxmg/4E8KMRpod8KT2EFB9ftG~1',
}

    census_headers = {
    'accept': 'application/json, text/plain, */*',
    'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
    'authorization': 'IFP_API ehi_visitor_id="C4667D4488814783A61F1BCFDA83C99C.ifp-c3po-app-prod-rs-v132-xbhcg"',
    'content-type': 'application/json;charset=UTF-8',
    'origin': 'https://www.ehealthinsurance.com',
    'pragma': 'no-cache',
    'priority': 'u=1, i',
    'referer': 'https://www.ehealthinsurance.com/small-business-health-insurance/group-health-insurance-plans?from=quote',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    }

    headers = {
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'authorization': 'IFP_API ehi_visitor_id="C4667D4488814783A61F1BCFDA83C99C.ifp-c3po-app-prod-rs-v132-xbhcg"',
        'priority': 'u=1, i',
        'referer': 'https://www.ehealthinsurance.com/smb-client/smbmedicalquote.html?groupId=7543132',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.plan_counter = 0
        self.read_file = self.read_xlsx_input_file()
        self.search_zip_codes = self.read_file.get('Input File')
        self.buisness_info = self.read_file.get('Business Info Input', {})[0]
        self.census_input = self.read_file.get('Census Input')
        self.output_file = self.read_file.get('Output File')
        self.cookie_jar = 1

    def parse_initial_requests(self, response):
        session_id = str(uuid4())
        print(session_id)
        url = 'https://www.ehealthinsurance.com/small-business-health-insurance/group-health-insurance-plans?fromPage=quote'
        # response.meta["zyte_api_automap"]  = {"geolocation": "US"}
        # response.meta["zyte_api_automap"]  = {"session": {"id": session_id}}
        yield Request(url, callback=self.parse_quote,
                      meta=response.meta, dont_filter=True)

    def parse_quote(self, response, **kwargs):
        # cookies, headers = self.get_cookies_headers(response)
        self.get_cookies_headers(response)
        auth_token = self.cookies.get('ehiVisitorId', '')
        alliance = self.cookies.get('ehi.alliance_id', '')
        group_id = '7543132'  # Default fallback if not in cookie

        # 2. Add Employee info and Zipcode wise search while "Percentage for employees" adjust in json_data
        update_census_url = 'https://www.ehealthinsurance.com/smb-api/smb/updateCensusPageInfo'

        zip_code = response.meta.get('zip_code', {})
        self.headers['authorization'] =  f'IFP_API ehi_visitor_id="{auth_token}"'
        json_data = self.get_json(auth_token, zip_code, alliance, group_id)
        # self.census_cookies['zipCode'] = str(zip_code.get('Zip Code', ''))
        self.cookies['zipCode'] = str(zip_code.get('Zip Code', ''))
        # response.meta["requestCookies"] = cookies
        response.meta["alliance"] = alliance
        response.meta["auth_token"] = auth_token
        response.meta["zip_code"] = zip_code
        response.meta["handle_httpstatus_all"] = True
        yield Request(url=update_census_url, method="POST", dont_filter=True, body=json.dumps(json_data),
                      headers=self.headers, cookies=self.cookies, callback=self.parse_update_census, meta=response.meta)

    def parse_update_census(self, response):
        try:
            group_id = response.json().get('groupId', '')
        except:
            group_id = '7543132'
            a=1

        timestamp = str(int(time.time() * 1000))
        alliance = response.meta.get('alliance', '') or 'ehe12862000'

        # Get All plans before step 11. Change “Percentage for dependents*” to 100%.
        plans_url = f"https://www.ehealthinsurance.com/smb-api/smb/{group_id}/plan/sg?includeRate=true&allianceId={alliance}&_={timestamp}"
        response.meta["group_id"] =  group_id
        yield Request(plans_url, headers=self.headers, callback=self.parse_plans, cookies=self.census_cookies, meta=response.meta, dont_filter=True)

    def parse_plans(self, response):
        group_id = response.meta.get('group_id', '')
        all_plans = response.json().get('plans', '')
        self.logger.info(f'{len(all_plans)} :: No of Plan Available')

        for plan in all_plans:
            try:
                carrier_id = plan.get('carrierId', '')
                plan_id = plan.get('planId', '')
                benefit_url = f'https://www.ehealthinsurance.com/smb-api/smb/{group_id}/{carrier_id}/{plan_id}/benefits?productLine=SG&fromEEPPE=false'
                plan_url = f'https://www.ehealthinsurance.com/smb-client/smb-plan-detail.html?carrierId={carrier_id}&planId={plan_id}'

                plan_headers = self.headers.copy()
                plan_headers['referer'] = response.meta.get('plan_url', '')
                response.meta["plan_info"] = plan
                response.meta["plan_url"] = plan_url

                yield Request(benefit_url, headers=plan_headers, meta=response.meta, callback=self.parse_plan, dont_filter=True)
            except Exception as e:
                self.logger.error(f'Plan: {response.meta.get('plan_info', {}).get('planName', '')} Face error :{e}')
                continue

    def parse_plan(self, response):
        plan_url = response.meta.get('plan_url', '')
        zip_code = response.meta.get('zip_code', {})
        plan = response.meta.get('plan_info', {})
        all_emp_plans = plan.get('employeeRateViews', [])
        all_premiums= ',\n '.join([str(emp.get('totalRate', 0.0)) for emp in all_emp_plans])
        premium = ceil(sum([emp.get('totalRate', 0.0) for emp in all_emp_plans]) / len(all_emp_plans)) if all_emp_plans else ''

        try:
            plan_detail = self.unescape_html(response.json())
        except Exception as e:
            plan_detail = {}
            self.logger.error(f"Plan processing failed: {e}")
            return

        try:
            benefits = plan_detail.get('benefits', [])
            get_benefit = lambda key: next(
                (b['benefitValue'] for b in benefits if b.get('benefitLabel') == key), '')
            g_drugs, p_drugs, n_p__drugs, s_drugs = self.get_tier_drugs(plan_detail)
            hgs_status = get_benefit('HSAEligible') or ''

            item = OrderedDict([
                ('Zip Code', zip_code.get('Zip Code', 0)),
                ('County', zip_code.get('County', '')),
                ('Effective Date', zip_code.get('Effective Date', '')),
                ('State', zip_code.get('State', '')),
                ('Competitor Name', plan_detail.get('planDetails', {}).get('carrierName', '')),
                ('Plan Name', plan_detail.get('planDetails', {}).get('planName', '')),
                ('Premium Per Employee', f'${premium}' if premium else ''),
                ('Percentage for dependents 0%', all_premiums),
                ('Plan type', plan_detail.get('planDetails', {}).get('categoryId', '')),
                ('Annual Deductible (per person)',
                 re.sub(r'[^\d]', '', get_benefit('AnnualDeductible').split()[0]) or ''),
                ('Out-of-Pocket limit',
                 re.sub(r'[^\d]', '', get_benefit('AnnualOutOfPocketLimit').split()[0]) or ''),
                ('Health Savings Account (HSA) eligible', 'Yes' if 'y' in hgs_status.lower() else 'No'),
                ('Primary Doctor', self.get_clean_benefit_value(get_benefit, 'PrimaryDoctor')),
                ('Specialist', self.get_clean_benefit_value(get_benefit, 'Specialist')),
                ('Generic/Tier 1 Drugs', g_drugs),
                ('Preferred Brand/Tier 2 Drugs', p_drugs),
                ('Non-Preferred Brand/Tier 3 Drugs:', n_p__drugs),
                ('Specialty/Tier 4 Drugs', s_drugs),
                ('Emergency room', self.get_clean_benefit_value(get_benefit, 'EmergencyRoom')),
                ('Urgent care', self.get_clean_benefit_value(get_benefit, 'UrgentCareFclt')),
                ('X-rays', self.get_clean_benefit_value(get_benefit, 'LabXRay')),
                ('CT,PET scans, MRIs',
                 self.get_clean_benefit_value(get_benefit, 'ImagingCTandPETscansMRIs')),
                ('Mental Health Outpatient Office Visit',
                 self.get_clean_benefit_value(get_benefit, 'OutpatientMentalHealth')),
                ('Psychiatric hospital stay',
                 self.get_clean_benefit_value(get_benefit, 'InpatientMentalHealth')),
                ('URL', plan_url),
            ])

            self.plan_counter += 1
            print(f'Plans Scraped :{self.plan_counter}')
            yield item
        except Exception as e:
            self.logger.error(f'Error In item yield Error:{e} &&&  URL:{plan_url}')

    def get_json(self, auth_token, zip_code, alliance, group_id):
        # Process phone number
        phone_no = self.buisness_info.get('Phone number', '')
        phone_str = str(phone_no).zfill(10) if isinstance(phone_no, int) else str(phone_no)
        area_num = phone_str[:3]
        phone_part = f"{phone_str[3:6]}-{phone_str[6:]}"

        # Basic business info
        zip_code_no = str(zip_code.get('Zip Code', ''))
        zip_code_county = zip_code.get('County', '')
        zip_code_st = zip_code.get('State', '')
        effective_date = zip_code.get('Effective Date', '')
        contact_first = self.buisness_info.get('First name', '')
        contact_last = self.buisness_info.get('Last name', '')
        email = self.buisness_info.get('Email address', '')
        num_employees = str(self.buisness_info.get('# of employees (including owner)', ''))

        # Generate employee list
        alpha = list(string.ascii_uppercase)
        employees = []
        id_counter = 1
        dependent_id = 1179123

        for i, person in enumerate(self.census_input):
            emp_name = alpha[i // 26] + alpha[i % 26]
            spouse = person.get('Include Spouse', '').strip().upper() == 'Y'
            dep_count = int(person.get('Dependent Count') or 0)

            emp = {
                'name': emp_name,
                'id': str(id_counter),
                'zipCode': zip_code_no,
                'employeeAge': str(person.get('Age', '')),
                'spouseAge': str(person.get('Spouse Age', '')),
                'numberOfChildren': str(dep_count)
            }

            if spouse:
                emp['addSpouse'] = 'true'

            if dep_count > 0:
                emp['dependents'] = []
                for j in range(dep_count):
                    age_key = 'Dependents Age' if j == 0 else f'Dependents Age_{j}'
                    dep_age = str(person.get(age_key, ''))
                    emp['dependents'].append({
                        'id': str(dependent_id),
                        'age': dep_age
                    })
                    dependent_id += 1

            employees.append(emp)
            id_counter += 1

        # Final payload
        json_data = {
            # 'ehiVisitorId': 'C4667D4488814783A61F1BCFDA83C99C.ifp-c3po-app-prod-rs-v132-xbhcg',
            'ehiVisitorId': auth_token,
            'groupId': '7543132',
            'allianceId': alliance if alliance else 'ehe12862000',
            'allianceSid': None,
            'alliancePid': None,
            'allianceCampaign': None,
            'performanceHorizonRefId': None,
            'quoteId': None,
            'fromProduct': 'SG',
            'clickId': '',
            'sicCode': '8999',
            'employeeEnrollingNumber': num_employees,
            'requestedEffectiveDate': effective_date,
            'stateAbbr': zip_code_st,
            'contactCompleted': 'true',
            'companyName': 'New Company',
            'zipCode': zip_code_no,
            'county': zip_code_county.upper(),
            'eeAmount': '100',
            'contactFirstName': contact_first,
            'contactLastName': contact_last,
            'email': email,
            'phoneExt': '',
            'employees': employees,
            'areaNum': area_num,
            'phone': phone_part
        }
        return json_data

    def read_xlsx_input_file(self):
        try:
            filename = glob.glob('ehealthinsurance_scraper/input/*xlsx')[0]

            if not filename:
                raise CloseSpider("Input file not found. Closing spider.")

            workbook = load_workbook(filename)
            all_data = {}

            for sheet in workbook.worksheets:
                sheet_name = sheet.title

                # Handle duplicate headers
                raw_headers = [cell.value for cell in sheet[1]]
                header_count = collections.Counter()
                headers = []

                for header in raw_headers:
                    if header_count[header] == 0:
                        headers.append(header)
                    else:
                        headers.append(f"{header}_{header_count[header]}")
                    header_count[header] += 1

                rows = []

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        # Replace None with ''
                        cleaned_row = [cell if cell is not None else '' for cell in row]

                        row_dict = dict(zip(headers, cleaned_row))

                        # Format Effective Date
                        if 'Effective Date' in row_dict and isinstance(row_dict['Effective Date'], datetime):
                            row_dict['Effective Date'] = row_dict['Effective Date'].strftime('%m/%d/%Y')

                        rows.append(row_dict)

                all_data[sheet_name] = rows

            return all_data

        except Exception as e:
            raise CloseSpider(f"An error occurred while reading the input file: {str(e)}. Closing spider.")

    def unescape_html(self, obj):
        if isinstance(obj, dict):
            return {k: self.unescape_html(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self.unescape_html(i) for i in obj]
        elif isinstance(obj, str):
            return html.unescape(obj)
        return obj

    def get_clean_benefit_value(self, get_benefit, label):
        try:
            value = ''.join(''.join(get_benefit(f'{label}').split('<br')[0:1]).split(':')[
                            1:]).strip() if '<br' in get_benefit(f'{label}') else get_benefit(
                f'{label}')
            return value
        except Exception as e:
            a = 2
            return ''

    def get_tier_drugs(self, plan_detail):
        try:
            g_drugs = p_drugs = n_p_drugs = s_drugs = ''

            # Get drug benefit value and normalize line breaks
            g_t_drugs_val = next((b['benefitValue'] for b in plan_detail.get('benefits', []) if
                                  b.get('benefitLabel') == 'RetailPrescriptionDrugs'), '').replace('&lt;br/&gt;',
                                                                                                   ', ')

            # Case 1: Label format like "Generic- $10 Preferred Brand- $20"
            if 'Non-Pref Brand' in g_t_drugs_val or 'Generic' in g_t_drugs_val:
                matches = re.findall(r'(Generic|Preferred Brand|Non-Pref Brand)[-\s]+(\$\d+)', g_t_drugs_val, re.I)
                label_map = {label.lower().replace(' ', '_'): value for label, value in matches}
                g_drugs = label_map.get('generic', '')
                p_drugs = label_map.get('preferred_brand', '')
                n_p_drugs = label_map.get('non-pref_brand', '')
                s_drugs = ''

            # Case 3: fallback if <br> with colon format
            elif '<br' in g_t_drugs_val and ':' in g_t_drugs_val:
                values = [
                    ''.join(t.split(':')[1:]).strip()
                    for t in g_t_drugs_val.split('<br') if ':' in t
                ]
                g_drugs, p_drugs, n_p_drugs, s_drugs = values

            # Case 1: Tier format like "Tier 1: $10, Tier 2: $20"
            elif 'Tier' in g_t_drugs_val:
                g_drugs, p_drugs, n_p_drugs, s_drugs = [''.join(tier.split(':')[1:]).strip() for tier in
                                                        g_t_drugs_val.split(',')]

            return g_drugs, p_drugs, n_p_drugs, s_drugs
        except Exception as e:
            return '', '', '', ''

    def closed(self, reason):
        logging.info(f"Spider closed. Total plans yielded: {self.plan_counter}")

    def get_cookies_headers(self, response):
        raw_headers = response.raw_api_response.get('httpResponseHeaders', [])

        # Step 1: Extract all headers except cookies
        headers = {}
        cookies = {}

        for item in raw_headers:
            name = item['name'].lower()
            value = item['value']

            if name == 'set-cookie':
                cookie_pair = value.split(';')[0]  # only take "key=value"
                if '=' in cookie_pair:
                    k, v = cookie_pair.split('=', 1)
                    self.cookies[k.strip()] = v.strip()
            else:

                self.headers[name] = value
        return cookies, headers

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(EhiSpiderSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        if self.search_zip_codes:
            zip_dict = self.search_zip_codes.pop()
            self.logger.info(f'{zip_dict.get('Zip Code', '')} Is Started Processing')
            self.crawler.engine.crawl(Request(url='https://books.toscrape.com', callback=self.parse_initial_requests,
                                              dont_filter=True, meta={'zip_code': zip_dict, 'handle_httpstatus_all':True, 'dont_merge_cookies': True}))
