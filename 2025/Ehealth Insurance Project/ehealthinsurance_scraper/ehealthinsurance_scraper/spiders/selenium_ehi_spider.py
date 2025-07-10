import io, re, gzip, json, html, glob
from math import ceil
from datetime import datetime
import time, logging, collections
from collections import OrderedDict
from http.cookies import SimpleCookie

from scrapy.exceptions import CloseSpider
from scrapy import Spider, Request, signals
from openpyxl.reader.excel import load_workbook

from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

# Silence all other loggers globally
logging.getLogger().setLevel(logging.CRITICAL)

# Specifically silence seleniumwire and urllib3 logs
logging.getLogger("seleniumwire").setLevel(logging.CRITICAL)
logging.getLogger("urllib3").setLevel(logging.CRITICAL)
logging.getLogger("selenium.webdriver.remote.remote_connection").setLevel(logging.CRITICAL)  # 🔇 hide selenium POST/DEBUG

# Your logger configuration
logger = logging.getLogger("ehealth_scraper")
logger.setLevel(logging.INFO)

class EhiSpiderSpider(Spider):
    name = "ehealth_scraper"
    allowed_domains = ["www.ehealthinsurance.com", "graph.ehealthinsurance.com"]

    custom_settings = {
        'OFFSITE_ENABLED': False,
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 408],
        'CONCURRENT_REQUESTS': 1,
        'FEED_EXPORTERS': {
            'xlsx': 'scrapy_xlsx.XlsxItemExporter',
        },
        'FEEDS': {
            f'outputs/Ehealth Records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx': {
                'format': 'xlsx',
                'encoding': 'utf-8',
                'fields': ['Zip Code', 'County', 'Effective Date', 'State', 'Competitor Name', 'Plan Name',
                           'Premium Per Employee', 'Plan type', 'Annual Deductible (per person)',
                           'Out-of-Pocket limit',
                           'Health Savings Account (HSA) eligible', 'Primary Doctor', 'Specialist',
                           'Generic/Tier 1 Drugs',
                           'Preferred Brand/Tier 2 Drugs', 'Non-Preferred Brand/Tier 3 Drugs:',
                           'Specialty/Tier 4 Drugs',
                           'Emergency room', 'Urgent care', 'X-rays', 'CT,PET scans, MRIs',
                           'Mental Health Outpatient Office Visit', 'Psychiatric hospital stay', 'URL']
            }
        },
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.plan_counter = 0
        self.read_file = self.read_xlsx_input_file()
        self.search_zip_codes = self.read_file.get('Input File')
        self.business_info = self.read_file.get('Business Info Input', {})[0]
        self.census_input = self.read_file.get('Census Input')
        self.output_file = self.read_file.get('Output File')
        self.driver = None

    def parse_initial_requests(self, response):
        zip_code = response.meta.get('zip_code', {})
        zip_code_str = str(zip_code.get('Zip Code', ''))
        self.driver = self.get_undetected_chrome_driver()
        company_form_submission = self.fill_company_form()
        if not company_form_submission:
            self.logger.error("❌ Company form submission failed.")
            return

        plans_data = self.input_employee_info(zip_code)
        if not plans_data:
            self.logger.warning(f"⚠️ No plan data returned for ZIP Code: {zip_code_str}")
            return

        # Normalize to list if response is a single dict
        if isinstance(plans_data, dict):
            plans_data = [plans_data]

        # Filter plans matching the ZIP code
        plans_list = [plan for plan in plans_data if isinstance(plan, dict) and plan.get('response', {}).get('quotePageExtend', {}).get('zipCode', '') == zip_code_str]

        if plans_list:
            plans = plans_list[0]  # If needed, use all with: `plans_list`
        else:
            time.sleep(5)
            plans = self.get_plans_dict(zip_code_str)

        group_id = ''.join(self.driver.current_url.split('groupId=')[1:])
        headers = plans.get('headers', {})
        cookies = plans.get('cookies', {})
        all_plans = plans.get('response', {}).get('plans', [])

        for plan in all_plans:
            carrier_id = plan.get('carrierId', '')
            plan_id = plan.get('planId', '')
            url= f'https://www.ehealthinsurance.com/smb-api/smb/{group_id}/{carrier_id}/{plan_id}/benefits?productLine=SG&fromEEPPE=false'
            plan_url = f'https://www.ehealthinsurance.com/smb-client/smb-plan-detail.html?carrierId={carrier_id}&planId={plan_id}'
            response.meta['plan'] = plan
            response.meta['zip_code'] = zip_code
            response.meta['plan_url'] = plan_url
            yield Request(url, headers=headers, meta=response.meta, callback=self.parse_plan, cookies=cookies, dont_filter=True)

    def parse_plan(self, response):
        plan_url = response.meta.get('plan_url', '')
        zip_code = response.meta.get('zip_code', {})
        plan = response.meta.get('plan', {})
        all_emp_plans = plan.get('employeeRateViews', [])
        premium = ceil(sum([emp.get('totalRate', 0.0) for emp in all_emp_plans]) / 15) if all_emp_plans else ''

        try:
            plan_detail = self.unescape_html(response.json())
        except Exception as e:
            self.logger.error(f"Plan processing failed: {e}")
            return

        try:
            benefits = plan_detail.get('benefits', [])
            get_benefit = lambda key: next(
                (b['benefitValue'] for b in benefits if b.get('benefitLabel') == key), '')
            g_drugs, p_drugs, n_p__drugs, s_drugs = self.get_tier_drugs(plan_detail)
            hgs_status = get_benefit('HSAEligible') or ''

            item = OrderedDict([
                ('Zip Code', zip_code.get('Zip Code', 0)),
                ('County', zip_code.get('County', '')),
                ('Effective Date', zip_code.get('Effective Date', '')),
                ('State', zip_code.get('State', '')),
                ('Competitor Name', plan_detail.get('planDetails', {}).get('carrierName', '')),
                ('Plan Name', plan_detail.get('planDetails', {}).get('planName', '')),
                ('Premium Per Employee', f'${premium}' if premium else ''),
                ('Plan type', plan_detail.get('planDetails', {}).get('categoryId', '')),
                ('Annual Deductible (per person)',
                 re.sub(r'[^\d]', '', get_benefit('AnnualDeductible').split()[0]) or ''),
                ('Out-of-Pocket limit',
                 re.sub(r'[^\d]', '', get_benefit('AnnualOutOfPocketLimit').split()[0]) or ''),
                ('Health Savings Account (HSA) eligible', 'Yes' if 'y' in hgs_status.lower() else 'No'),
                ('Primary Doctor', self.get_clean_benefit_value(get_benefit, 'PrimaryDoctor')),
                ('Specialist', self.get_clean_benefit_value(get_benefit, 'Specialist')),
                ('Generic/Tier 1 Drugs', g_drugs),
                ('Preferred Brand/Tier 2 Drugs', p_drugs),
                ('Non-Preferred Brand/Tier 3 Drugs:', n_p__drugs),
                ('Specialty/Tier 4 Drugs', s_drugs),
                ('Emergency room', self.get_clean_benefit_value(get_benefit, 'EmergencyRoom')),
                ('Urgent care', self.get_clean_benefit_value(get_benefit, 'UrgentCareFclt')),
                ('X-rays', self.get_clean_benefit_value(get_benefit, 'LabXRay')),
                ('CT,PET scans, MRIs', self.get_clean_benefit_value(get_benefit, 'ImagingCTandPETscansMRIs')),
                ('Mental Health Outpatient Office Visit',
                 self.get_clean_benefit_value(get_benefit, 'OutpatientMentalHealth')),
                ('Psychiatric hospital stay',
                 self.get_clean_benefit_value(get_benefit, 'InpatientMentalHealth')),
                ('URL', plan_url),
             ])

            self.plan_counter += 1
            print(f'Plans Scraped :{self.plan_counter}')
            yield item
        except Exception as e:
            self.logger.error(f'Error In item yield Error:{e} &&&  URL:{plan_url}')

    def read_xlsx_input_file(self):
        try:
            filename = glob.glob('ehealthinsurance_scraper/input/*xlsx')[0]

            if not filename:
                raise CloseSpider("Input file not found. Closing spider.")

            workbook = load_workbook(filename)
            all_data = {}

            for sheet in workbook.worksheets:
                sheet_name = sheet.title

                # Handle duplicate headers
                raw_headers = [cell.value for cell in sheet[1]]
                header_count = collections.Counter()
                headers = []

                for header in raw_headers:
                    if header_count[header] == 0:
                        headers.append(header)
                    else:
                        headers.append(f"{header}_{header_count[header]}")
                    header_count[header] += 1

                rows = []

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        # Replace None with ''
                        cleaned_row = [cell if cell is not None else '' for cell in row]

                        row_dict = dict(zip(headers, cleaned_row))

                        # Format Effective Date
                        if 'Effective Date' in row_dict and isinstance(row_dict['Effective Date'], datetime):
                            row_dict['Effective Date'] = row_dict['Effective Date'].strftime('%m/%d/%Y')

                        rows.append(row_dict)

                all_data[sheet_name] = rows

            return all_data

        except Exception as e:
            raise CloseSpider(f"An error occurred while reading the input file: {str(e)}. Closing spider.")

    def unescape_html(self, obj):
        if isinstance(obj, dict):
            return {k: self.unescape_html(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self.unescape_html(i) for i in obj]
        elif isinstance(obj, str):
            return html.unescape(obj)
        return obj

    def get_clean_benefit_value(self, get_benefit, label):
        try:
            value = ''.join(''.join(get_benefit(f'{label}').split('<br')[0:1]).split(':')[
                            1:]).strip() if '<br' in get_benefit(f'{label}') else get_benefit(
                f'{label}')
            return value
        except Exception as e:
            a = 2
            return ''

    def get_tier_drugs(self, plan_detail):
        try:
            g_drugs = p_drugs = n_p_drugs = s_drugs = ''

            # Get drug benefit value and normalize line breaks
            g_t_drugs_val = next((b['benefitValue'] for b in plan_detail.get('benefits', []) if
                                  b.get('benefitLabel') == 'RetailPrescriptionDrugs'), '').replace('&lt;br/&gt;',
                                                                                                   ', ')

            # Case 1: Label format like "Generic- $10 Preferred Brand- $20"
            if 'Non-Pref Brand' in g_t_drugs_val or 'Generic' in g_t_drugs_val:
                matches = re.findall(r'(Generic|Preferred Brand|Non-Pref Brand)[-\s]+(\$\d+)', g_t_drugs_val, re.I)
                label_map = {label.lower().replace(' ', '_'): value for label, value in matches}
                g_drugs = label_map.get('generic', '')
                p_drugs = label_map.get('preferred_brand', '')
                n_p_drugs = label_map.get('non-pref_brand', '')
                s_drugs = ''

            # Case 3: fallback if <br> with colon format
            elif '<br' in g_t_drugs_val and ':' in g_t_drugs_val:
                values = [
                    ''.join(t.split(':')[1:]).strip()
                    for t in g_t_drugs_val.split('<br') if ':' in t
                ]
                g_drugs, p_drugs, n_p_drugs, s_drugs = values

            # Case 1: Tier format like "Tier 1: $10, Tier 2: $20"
            elif 'Tier' in g_t_drugs_val:
                g_drugs, p_drugs, n_p_drugs, s_drugs = [''.join(tier.split(':')[1:]).strip() for tier in
                                                        g_t_drugs_val.split(',')]

            return g_drugs, p_drugs, n_p_drugs, s_drugs
        except Exception as e:
            return '', '', '', ''

    def closed(spider, reason):
        try:
            if spider.driver:
                spider.driver.quit()  # shuts down the entire driver process
                logging.info("WebDriver closed successfully.")
        except Exception as e:
            logging.warning(f"Error closing WebDriver: {e}")

        logging.info(f"Spider closed. Total plans yielded: {spider.plan_counter}")

    def get_undetected_chrome_driver(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])  # ⛔ suppress ChromeDriver logs

        s_options = {
            'connection_timeout': 500,
            'timeout': 500,
        }

        driver = webdriver.Chrome(seleniumwire_options=s_options, options=options)
        return driver

    def fill_company_form(self):
        url = 'https://www.ehealthinsurance.com/small-business-health-insurance/group-health-insurance-plans?fromPage=quote'
        self.logger.info(f"Navigating to: {url}")

        try:
            self.driver.get(url)
            form_text = 'Get Group Health Insurance Quote'
            WebDriverWait(self.driver, 100).until(
                lambda d: (d.execute_script("return document.readyState") == "complete" or
                           form_text in d.page_source))

            # Data extraction
            first_name = self.business_info.get('First name', '')
            last_name = self.business_info.get('Last name', '')
            email_address = self.business_info.get('Email address', '')
            zip_code = self.business_info.get('ZIP code', '')
            num_employees = str(self.business_info.get('# of employees (including owner)', ''))
            business_phone = self.business_info.get('Phone number', '')

            # Fill form fields
            def fill_input(by, value, text):
                elem = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((by, value)))
                elem.clear()
                elem.send_keys(text)
                self.logger.info(f"Filled field {value} with: {text}")

            fill_input(By.ID, 'firstName', first_name)
            fill_input(By.ID, 'lastName', last_name)
            fill_input(By.ID, 'email', email_address)
            fill_input(By.ID, 'zip', zip_code)
            fill_input(By.ID, 'phone', business_phone)

            # Select number of employees
            select_elem = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.ID, 'employeeEnrollingNumber')))
            Select(select_elem).select_by_value(num_employees)
            self.logger.info(f"Selected number of employees: {num_employees}")

            # Submit form
            submit_btn = WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.ID, 'submitToQuote')))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", submit_btn)
            submit_btn.click()
            self.logger.info("Clicked submit button")

            # Wait for listing page
            listing_page_text = 'Find the best plan for your business'
            WebDriverWait(self.driver, 100).until(
                lambda d: (d.execute_script("return document.readyState") == "complete" or
                           listing_page_text in d.page_source))

            self.logger.info("Successfully reached listing page")
            return True

        except Exception as e:
            self.logger.error(f"Error during form fill or navigation: {e}")
            return False

    def input_employee_info(self, zip_code):
        try:
            # Get zip & effective date
            zip_code_no = str(zip_code.get('Zip Code', ''))
            effective_date = zip_code.get('Effective Date', '')

            self.logger.info("▶️ Starting to input employee info...")

            WebDriverWait(self.driver, 100).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.census-notification a.edit-census')))
            edit_link = self.driver.find_element(By.CSS_SELECTOR, 'div.census-notification a.edit-census')
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", edit_link)

            try:
                edit_link.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", edit_link)
            self.logger.info("✔️ Clicked 'edit your employee details'")

            # Fill zip code
            zip_elem = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.ID, 'zip')))
            zip_elem.clear()
            zip_elem.send_keys(zip_code_no)

            # Set employer contribution to 100%
            Select(WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.ID, 'employerContribution'))
            )).select_by_value("100")

            self.logger.info("✔️ Set employer contribution to 100%")

            # Set Effective Date (remove readonly first)
            date_input = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.ID, 'requestEffectiveDate')))
            self.driver.execute_script("arguments[0].removeAttribute('readonly')", date_input)
            date_input.clear()
            date_input.send_keys(effective_date)
            self.logger.info(f"✔️ Set effective date to: {effective_date}")

            # Click "Add Employee Information"
            add_emp_btn = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID, 'showEmployeeInfo')))
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", add_emp_btn)
            try:
                add_emp_btn.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", add_emp_btn)
            self.logger.info("✔️ Clicked 'Add Employee Information'")
            time.sleep(3)
            # Fill employee census details
            self.fill_employee_census(zip_code_no)

            # Click "View Quotes"
            view_quotes_btn = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID, 'submitToQuote')))
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", view_quotes_btn)
            self.driver.execute_script("arguments[0].click();", view_quotes_btn)
            self.logger.info("✔️ Clicked 'View Quotes' button")

            # Wait until the quote page fully loads
            WebDriverWait(self.driver, 150).until(
                lambda d: d.execute_script("return document.readyState") == "complete")

            time.sleep(15)
            self.logger.info("✅ Quote page loaded successfully")

            data = []
            # Find request with the matching URL
            for request in self.driver.requests:
                if request.response and "sg?includeRate=true&allianceId" in request.url:
                    try:
                        headers = dict(request.headers)
                        raw_cookie_header = headers.get('cookie', '')
                        cookie = SimpleCookie()
                        cookie.load(raw_cookie_header)
                        cookies_dict = {key: morsel.value for key, morsel in cookie.items()}

                        raw_body = request.response.body
                        with gzip.GzipFile(fileobj=io.BytesIO(raw_body)) as f:
                            decompressed = f.read().decode('utf-8', errors='replace')
                        response_data  = json.loads(decompressed)

                        data.append({
                            'url': request.url,
                            'headers': headers,
                            'cookies': cookies_dict,
                            'response': response_data
                        })
                        time.sleep(2)
                    except Exception as e:
                        a=2

            if data:
                return data

            self.logger.warning("⚠️ No matching alliance request captured")
            return None

        except Exception as e:
            self.logger.error(f"❌ Failed in employee info input: {e}")
            return None

    def is_block(self):
        block_text = 'Access Denied' or 'You don\'t have permission to access'
        if block_text in self.driver.page_source:
            self.logger.warning(f'Website is Not accessible due to Block Please retry with Vpn')
            # self.driver.close()
            self.logger.info(f'Now Proceed for the next Zip code')
            self.spider_idle()

    def fill_employee_census(self,zip_code_no):
        i = 0
        try:
            for i, emp in enumerate(self.census_input[:15]):
                age = str(emp.get("Age", ""))
                spouse = emp.get("Include Spouse", "N") == "Y"
                spouse_age = str(emp.get("Spouse Age", "") or "")
                dependents = str(emp.get("Dependent Count", "") or "0")

                # ✅ Fill ZIP code first
                zip_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, f'emp-zip{i}'))
                )
                zip_input.clear()
                zip_input.send_keys(zip_code_no)
                self.logger.info(f"Set ZIP code for Employee {i + 1}: {zip_code_no}")

                # Fill Age
                age_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, f'emp-age{i}'))
                )
                age_input.clear()
                age_input.send_keys(age)

                # Handle Spouse
                if spouse:
                    spouse_checkbox = self.driver.find_element(By.ID, f'emp-spouse{i}')
                    if not spouse_checkbox.is_selected():
                        self.driver.execute_script("arguments[0].click();", spouse_checkbox)

                    # Wait for spouse age field and enter age
                    spouse_age_input = WebDriverWait(self.driver, 5).until(
                        EC.visibility_of_element_located((By.ID, f'emp-spouse-age{i}'))
                    )
                    spouse_age_input.clear()
                    spouse_age_input.send_keys(spouse_age)

                # Set number of dependents
                dependents_select = Select(WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, f'emp-dependents{i}'))
                ))
                dependents_select.select_by_value(dependents)

                self.logger.info(
                    f"Filled Employee {i + 1}: Age={age}, Spouse={spouse}, Spouse Age={spouse_age}, Dependents={dependents}")
        except Exception as e:
            self.logger.error(f"[Employee Census] Error at Employee {i + 1}: {e}")

    def get_plans_dict(self, zip_code_str):
        data = []
        # Find request with the matching URL
        for request in self.driver.requests:
            if request.response and "sg?includeRate=true&allianceId" in request.url:
                try:
                    headers = dict(request.headers)
                    raw_cookie_header = headers.get('cookie', '')
                    cookie = SimpleCookie()
                    cookie.load(raw_cookie_header)
                    cookies_dict = {key: morsel.value for key, morsel in cookie.items()}

                    raw_body = request.response.body
                    with gzip.GzipFile(fileobj=io.BytesIO(raw_body)) as f:
                        decompressed = f.read().decode('utf-8', errors='replace')
                    response_data = json.loads(decompressed)

                    data.append({
                        'url': request.url,
                        'headers': headers,
                        'cookies': cookies_dict,
                        'response': response_data
                    })
                    time.sleep(2)
                except Exception as e:
                    a = 2

        plans_list = [plan for plan in data if
                      isinstance(plan, dict) and plan.get('response', {}).get('quotePageExtend', {}).get('zipCode',
                                                                                                         '') == zip_code_str]

        if plans_list:
            plans = plans_list[0]
            return plans

        return {}

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(EhiSpiderSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        if self.search_zip_codes:

            if self.driver:
                try:
                    self.logger.info("Closing existing driver before opening new one.")
                    self.driver.quit()
                except Exception as e:
                    self.logger.warning(f"Error while closing driver: {e}")
                finally:
                    self.driver = None  # Always reset it

            zip_dict = self.search_zip_codes.pop()
            self.logger.info(f'{zip_dict.get('Zip Code', '')} Is Started Processing')
            self.crawler.engine.crawl(Request(url='https://books.toscrape.com', callback=self.parse_initial_requests,
                                              dont_filter=True, meta={'zip_code': zip_dict, 'handle_httpstatus_all':True, 'dont_merge_cookies': True}))

