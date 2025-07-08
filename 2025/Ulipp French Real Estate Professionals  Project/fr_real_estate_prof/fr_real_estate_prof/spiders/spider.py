import csv
import glob
import html
import os
from copy import deepcopy
from datetime import datetime
from urllib.parse import urljoin
from collections import OrderedDict

import unicodedata
from openpyxl.reader.excel import load_workbook
from scrapy import Spider, Request, FormRequest, Selector

from twocaptcha import TwoCaptcha

api_key = '907077086f4e4873a6c8d3589e8176ed'
API_KEY = '907077086f4e4873a6c8d3589e8176ed'


class FrrealestateSpider(Spider):
    name = "Real_Estate"
    allowed_domains = ["registre.oaciq.com"]
    start_urls = ["https://registre.oaciq.com/fr/trouver-courtier"]

    custom_settings = {
        'DOWNLOADER_CLIENT_TLS_VERIFY': False,
        'OFFSITE_ENABLED': False,

        'RETRY_TIMES': 10,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408, 429],

        # 'FEED_EXPORTERS': {
        #     'xlsx': 'scrapy_xlsx.XlsxItemExporter',
        # },
        'FEEDS': {
            f'outputs/URLS_French Real Estate Professional Records_{datetime.now().strftime("%d%m%Y%H%M")}.csv': {
                'format': 'csv',
                'fields': ['Full Name', 'Email Address', 'Contact', 'Address', 'URL'],
            }
        },
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
       }

    pagination_headers = {
    'Accept': '*/*;q=0.5, text/javascript, application/javascript, application/ecmascript, application/x-ecmascript',
    'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Referer': '',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
    'X-CSRF-Token': '',
    'X-Requested-With': 'XMLHttpRequest',
   }

    headers_1 = {
        'Accept': '*/*;q=0.5, text/javascript, application/javascript, application/ecmascript, application/x-ecmascript',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Origin': 'https://registre.oaciq.com',
        'Referer': 'https://registre.oaciq.com/en/find-broker/C6E365F4',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
        'X-CSRF-Token': 'tYID3/71yplYTHrBQR7QykdjbIRb26Q0Zz4Hrn+6JnFA8w3NzDh0KaqX8pH56E4I3MLywRjxe/El2oRSHWl9ig==',
        'X-Requested-With': 'XMLHttpRequest',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.item_scraped = 0
        os.makedirs('outputs', exist_ok=True)
        self.previous_records = self.read_previous_records()
        self.csv_file_path = f'outputs/French Real Estate Professional Records_Live_{datetime.now().strftime("%d%m%Y%H%M")}.csv'

    def start_requests(self):
        yield Request(url=self.start_urls[0], headers=self.headers)

    def parse(self, response, **kwargs):
        token = response.css('[name="csrf-token"] ::attr(content)').get('')
        self.headers['Content-Type'] = 'application/x-www-form-urlencoded'
        self.headers['Origin'] = 'https://registre.oaciq.com'
        self.headers['Referer'] = self.start_urls[0]
        data = self.form_data(token)

        yield FormRequest(self.start_urls[0], formdata=data, headers=self.headers,
                        method='POST', callback=self.parse_listings, dont_filter=True)

    def parse_listings(self, response):
        brokers = response.css('#find-brokers-result tbody tr')

        for broker in brokers:
            b_url = broker.css('a::attr(href)').get('')
            url = urljoin(response.url, b_url)

            #test
            item = OrderedDict()
            name = broker.css('.fa-user + a::text, .fa-home + a::text').get('')
            if not name:
                a=1

            item['Full Name'] = self.normalize_text(name)
            item['URL'] = url
            yield item

        next_page = response.css('.next ::attr(href)').get('')
        if next_page:
            url = urljoin(response.url, next_page)
            token = response.css('[name="csrf-token"] ::attr(content)').get('')
            self.pagination_headers['X-CSRF-Token'] = token
            yield Request(url,  headers=self.headers, callback=self.parse_listings)

    def parse_broker(self, response):
        broker_name = response.css('#find-broker h1::text').get('').strip()
        solver = TwoCaptcha(api_key)
        site_key = response.css('[data-turnstile-site-key-value] ::attr(data-turnstile-site-key-value)').get(
            '') or '0x4AAAAAAA0B6y-PiDlytMN0'
        url_path = response.css('#url_path::attr(value) ').get('')
        result = solver.turnstile(
            sitekey=site_key,
            url=response.url,
        )

        url = "https://registre.oaciq.com/show_info_broker?locale=en"
        data = {
            'utf8': '✓',
            'cf-turnstile-response': result.get('code'),
            'url_path': url_path,
        }

        headers_1 = deepcopy(self.headers_1)
        headers_1['X-CSRF-Token'] = response.css('[name="csrf-token"]::attr(content)').get('')
        response.meta['broker_name'] = broker_name
        response.meta['url'] = response.url
        yield FormRequest(url=url, headers=headers_1, callback=self.parse_broker_detail,
                          meta=response.meta, formdata=data, priority=1000)

    def parse_broker_detail(self, response):
        sel = Selector(text=response.text.replace('\\n', '').replace('\\"', '"').replace("\\'", "'").replace('\\\\', '\\').replace('\\/', '/').replace('\\', ''))
        raw_html = response.text.split('$("#find_broker_info").html("')[1].split('$("form.contact").remove();')[0]
        cleaned = raw_html.replace('\\n', '').replace('\\"', '"').replace("\\'", "'").replace('\\\\', '\\').replace('\\/', '/').replace('\\', '')
        resp = Selector(text=cleaned)

        business_address = ''.join(resp.css('td:contains("Business address") + td::text').getall()) or  ' '.join(sel.xpath('//tr[td[1][contains(text(), "Business address")]]/td[2]//text()').getall()).strip()
        telephone = ''.join(resp.css('td:contains("Telephone") + td::text').getall()) or ''.join(resp.css('td:contains("Téléphone") + td::text').getall()) or sel.xpath('//tr[td[1][contains(text(), "Telephone")]]/td[2]/text()').get('')
        email = resp.css('td:contains("@") ::text').get('')

        item = OrderedDict()
        item['Full Name'] = self.normalize_text(response.meta.get('broker_name', ''))
        item['Email Address'] = email
        item['Contact'] = telephone
        item['Address'] = self.normalize_text(business_address)
        item['URL'] = response.meta.get('url', '')

        if item['Email Address'] or item['Contact']:
            self.item_scraped += 1
            print('Brokers Are Scraped :', self.item_scraped)
            yield item
        else:
            a=1

    def form_data(self, token):
        data = {
            'utf8': '✓',
            'authenticity_token': token,
            'find_broker[name]': '',
            'find_broker[licence_number]': '',
            'find_broker[include_revoked_brokers]': '0',
            'find_broker[area_of_practice]': '',
            'find_broker[agency_name]': '',
            'find_broker[region]': '',
            'find_broker[city]': '',
            'commit': 'Rechercher',
        }
        return data

    def write_csv(self, item):
        fieldnames = ['Full Name', 'Email Address', 'Contact', 'Address', 'URL']
        with open(self.csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            if csvfile.tell() == 0:
                writer.writeheader()
            writer.writerow({k: self.normalize_text(item.get(k, '')) for k in fieldnames})


    def normalize_text(self, text):
        if not isinstance(text, str):
            return text

        # Remove HTML entities and unwanted escape characters
        text = html.unescape(text)
        text = text.replace('\xa0', ' ')  # non-breaking space
        text = text.replace('\\"', '"').replace("\\'", "'")
        text = text.replace('\\\\', '\\')
        text = text.replace('\\n', '\n').replace('\\/', '/')

        # Normalize unicode while preserving accents (NFC form keeps é, è, ç etc.)
        text = unicodedata.normalize('NFC', text)

        return text.strip()

    def read_previous_records(self):
        records = {}
        try:
            filenames = glob.glob('outputs/French*.xlsx')  # more flexible
            if not filenames:
                print("[INFO] No previous Excel files found.")
                return {}

            for filepath in filenames:
                print(f"[INFO] Reading: {filepath}")
                workbook = load_workbook(filepath)
                sheet = workbook.active
                headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
                print(f"[DEBUG] Headers: {headers}")

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    url = row_data.get('URL') or row_data.get('url')
                    if url and url not in records:
                        records[url] = row_data

        except Exception as e:
            print(f"[ERROR] Failed to read records: {e}")

        return records