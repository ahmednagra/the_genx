import csv
import glob
import html
import os
from collections import OrderedDict
from copy import deepcopy
from datetime import datetime
from typing import Iterable

import scrapy
import unicodedata
from openpyxl.reader.excel import load_workbook
from scrapy import Request, FormRequest, Selector
from twocaptcha import TwoCaptcha


class OaciqSpiderSpider(scrapy.Spider):
    name = "oaciq_spider"
    custom_settings = {
        'DOWNLOADER_CLIENT_TLS_VERIFY': False,
        'OFFSITE_ENABLED': False,

        'RETRY_TIMES': 10,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408, 429],

        'FEED_EXPORTERS': {
            'xlsx': 'scrapy_xlsx.XlsxItemExporter',
        },
        'FEEDS': {
            f'outputs/French Real Estate Professional Records_{datetime.now().strftime("%d%m%Y%H%M")}.xlsx': {
            'format': 'xlsx',
            'fields': ['Full Name', 'Email Address', 'Contact', 'Address', 'URL'],
            }
        },
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
        'Referer': 'https://registre.oaciq.com/en/find-broker',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"',
        # 'Cookie': 'didomi_token=eyJ1c2VyX2lkIjoiMTk3NjcwYWEtODU4YS02OTM2LTkwN2EtNTFhYTQ0YzU5YzYzIiwiY3JlYXRlZCI6IjIwMjUtMDYtMTNUMDI6MDc6MjEuMTc2WiIsInVwZGF0ZWQiOiIyMDI1LTA2LTEzVDAyOjA3OjIzLjMxOVoiLCJ2ZW5kb3JzIjp7ImVuYWJsZWQiOlsiZ29vZ2xlIiwiYzpnb29nbGVhbmEtNFRYbkppZ1IiLCJjOmZhY2Vib29rLVd0Q0JDQ3dwIl19LCJ2ZW5kb3JzX2xpIjp7ImVuYWJsZWQiOlsiZ29vZ2xlIiwiYzphdXRoZW50aWZpLTNRVEtENFByIl19LCJ2ZXJzaW9uIjoyLCJhYyI6IkFGbUFDQUZrLkFGbUFDQUZrIn0=; euconsent-v2=CQS8iQAQS8iQAAHABBENBuFgAP_AAABAAAAAGSQAgF5gMkAySAEAvMBkgAhQAASgAwABBaQdABgACC0gSADAAEFpAAAA.f_gAAAgAAAAA; _gcl_au=1.1.420280246.1749780443; _ga=GA1.1.1950972276.1749780444; _hjSessionUser_47208=eyJpZCI6IjY1OThhN2NlLWM5MmItNWU0NS05MjdhLWFiYmI2YTViZWRhMSIsImNyZWF0ZWQiOjE3NDk3ODA0NDI2NzIsImV4aXN0aW5nIjp0cnVlfQ==; _ga_SJ35FCRYGZ=GS2.1.s1749783767$o2$g1$t1749783790$j38$l0$h0; _oaciq_session_dksfhkdsfhdkdfhs=U0VLVUplNm52aUJTNTNxOEJHckNkYmV0TmRZZjcxM2NxYStCTkhYNCsrbzhST2pMZmZwRWtPQk9wbUVoQkFOV2NCdjcwSWRRZDJkOE54bzU2ZDJKTVlnVVRUTHNubEtILzRXQXBwOG0vUFBrMDFFaksvOTl5dlMySCtzMDJSNnJWdVBwN0ZtWVdnME5TbmlSeWZYYTZBMFE4am51NUJXUmpQRS9Kc0lOK3JwRDF4RFF1NUlqdHdrSHdNYjM1bUxISzl4WGVoRHM4WURTMllwVnpGNVYxSDRTRFZ3QW8yU1BjVE5CbkowV0lqRk1McjgvT2plS0RZVHRldnF0TXJ4OC0tOXVwWEllTkRxdzFOMkdONUR0a3Bhdz09--81ec3511b2abd24a4e3c9a078fc3424552d8b4c5',
    }
    headers_1 = {
        'Accept': '*/*;q=0.5, text/javascript, application/javascript, application/ecmascript, application/x-ecmascript',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Origin': 'https://registre.oaciq.com',
        'Pragma': 'no-cache',
        'Referer': 'https://registre.oaciq.com/en/find-broker/C6E365F4',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
        'X-CSRF-Token': 'tYID3/71yplYTHrBQR7QykdjbIRb26Q0Zz4Hrn+6JnFA8w3NzDh0KaqX8pH56E4I3MLywRjxe/El2oRSHWl9ig==',
        'X-Requested-With': 'XMLHttpRequest',
        'sec-ch-ua': '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"',
        # 'Cookie': '_hjSession_47208=eyJpZCI6ImEyYjBkMmNlLTM3NDUtNDVhMC1hOWE1LTllODE0ODQ2NTFiYiIsImMiOjE3NDk3ODA0NDI2NzMsInMiOjAsInIiOjAsInNiIjowLCJzciI6MCwic2UiOjAsImZzIjoxLCJzcCI6MH0=; didomi_token=eyJ1c2VyX2lkIjoiMTk3NjcwYWEtODU4YS02OTM2LTkwN2EtNTFhYTQ0YzU5YzYzIiwiY3JlYXRlZCI6IjIwMjUtMDYtMTNUMDI6MDc6MjEuMTc2WiIsInVwZGF0ZWQiOiIyMDI1LTA2LTEzVDAyOjA3OjIzLjMxOVoiLCJ2ZW5kb3JzIjp7ImVuYWJsZWQiOlsiZ29vZ2xlIiwiYzpnb29nbGVhbmEtNFRYbkppZ1IiLCJjOmZhY2Vib29rLVd0Q0JDQ3dwIl19LCJ2ZW5kb3JzX2xpIjp7ImVuYWJsZWQiOlsiZ29vZ2xlIiwiYzphdXRoZW50aWZpLTNRVEtENFByIl19LCJ2ZXJzaW9uIjoyLCJhYyI6IkFGbUFDQUZrLkFGbUFDQUZrIn0=; euconsent-v2=CQS8iQAQS8iQAAHABBENBuFgAP_AAABAAAAAGSQAgF5gMkAySAEAvMBkgAhQAASgAwABBaQdABgACC0gSADAAEFpAAAA.f_gAAAgAAAAA; _gcl_au=1.1.420280246.1749780443; _ga=GA1.1.1950972276.1749780444; _ga_SJ35FCRYGZ=GS2.1.s1749780443$o1$g0$t1749780445$j58$l0$h646584700; _hjSessionUser_47208=eyJpZCI6IjY1OThhN2NlLWM5MmItNWU0NS05MjdhLWFiYmI2YTViZWRhMSIsImNyZWF0ZWQiOjE3NDk3ODA0NDI2NzIsImV4aXN0aW5nIjp0cnVlfQ==; _oaciq_session_dksfhkdsfhdkdfhs=UTJIVXZjbEJiQmNrbGZCUDAzRHk3dFcrSHh5aCtkeU5waTFWNEdOZ0h5RTNzMjd2dnA4eDQvKzJMSWpGTjB4T21VbFZwZmExakRoeVBTNTIvUnc1ZHhyYWNWUG41Y0ZsdFVKTWtaQWl3cUlSUFNLR1IzU25oS0w3K1RTbkk3MnBlNlFreTNWT3RsdUs1N1VscTAwd0xpL0Jqamt4N3cxSGtwVHFoUHBWNXo2NTN2ckszSzFnN3BrOTVlcWFGbG5iZFR3MUJOeTUxR0ZpVTRsajNnM3NQcHlzcjZWMitqYmp1WHhDSjN2TDhPcUNwN3dFc0E1UmpYUUYrS1VFREtVNi0tNmFIZ1JVU0hnZ1h0UmtDLzdZUnJZQT09--cf2007eb5202f7258ac201ca2a3f747113534104; _oaciq_session_dksfhkdsfhdkdfhs=ajFMRjZOaGNLMCtKMWFPQXVWbjJSRjFQS01mY0xNTlhJZ2V4UVBTSGV6OVhidVNaUHc4U3pSelNaMys0cWtrTlI2QkNmZnordzdrZHhPbzgrSGxibitROHk0NDF5TndHQlh1eHFmaFNHZ01WcUlVbHhtMTBrd2ZoZ1ZYejRaZ3lpcXM0WG5jL0czSGprY0U5dURkYzlKTThLYzZvRUNDK0h2ZGN2alo0NExBenA1U1ZwQVhZT2hkOE1pRW9iNDA3bWtEb3NqRlIvUE5qVUdmQk5QUVIwUWt0MUkwQVhtQWI4V2x3TXU4YXJuRHpOK3JjNDhLV1cwRG02WC9NWFJkSzhJTzIyaHdMbTAzRVJYWDQwRTZQZENwVExpQVhBSHMrZ2ZRWUErR1dSM0pMamkycTNXYWkrWW14cFRmYlZDbVRzS3Raem1sUUU2NHFvZkFERWlENUw1SGZKdmUvWWYzcWVkR2FreEJjMFFqMWF2aFBQM00zUDBZc1JzeDB5QjZYLS1JeFNiU2tJSFpRNmNZUU1QNWQ0Z2RBPT0%3D--6205a28d266aa7d038b81e1ad9a87d20fa11757a'
    }
    cookie_jar = 1

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.item_scraped = 0
        os.makedirs('outputs', exist_ok=True)
        self.previous_records = self.read_previous_records()

    def start_requests(self) -> Iterable[Request]:
        csv_file_path = glob.glob('URLS*.csv')[0]
        urls = []

        with open(csv_file_path, mode='r', encoding='utf-8') as f:
            urls.extend(csv.DictReader(f))

        for row in urls:
            url = row.get('URL', '')
            if url:
                self.cookie_jar = self.cookie_jar + 1
                yield scrapy.Request(url,
                    # url='https://registre.oaciq.com/en/find-broker/C6E365F4',
                    headers=self.headers,
                    meta={'cookiejar': self.cookie_jar, 'row':row}
                )

    def parse(self, response, **kwargs):
        url_path = response.css('#url_path::attr(value) ').get('')
        api_key = '907077086f4e4873a6c8d3589e8176ed'
        solver = TwoCaptcha(api_key)

        result = solver.turnstile(
            sitekey='0x4AAAAAAA0B6y-PiDlytMN0',
            url=response.url,
        )

        url = "https://registre.oaciq.com/show_info_broker?locale=en"
        data = {
            'utf8': '✓',
            'cf-turnstile-response': result.get('code'),
            'url_path': url_path,
        }
        headers_1 = deepcopy(self.headers_1)
        headers_1['X-CSRF-Token'] = response.css('[name="csrf-token"]::attr(content)').get('')
        yield FormRequest(url, formdata=data, headers=headers_1,
            callback=self.parse_detail, meta=response.meta, priority=1000
        )

    def parse_detail(self, response):
        sel = Selector(text=response.text.replace('\\n', '').replace('\\"', '"').replace("\\'", "'").replace('\\\\', '\\').replace(
                '\\/', '/').replace('\\', ''))

        raw_html = response.text.split('$("#find_broker_info").html("')[1].split('$("form.contact").remove();')[0]
        cleaned = raw_html.replace('\\n', '').replace('\\"', '"').replace("\\'", "'").replace('\\\\', '\\').replace(
            '\\/', '/').replace('\\', '')
        resp = Selector(text=cleaned)

        business_address = ''.join(resp.css('td:contains("Business address") + td::text').getall()) or ' '.join(
            sel.xpath('//tr[td[1][contains(text(), "Business address")]]/td[2]//text()').getall()).strip()
        telephone = ''.join(resp.css('td:contains("Telephone") + td::text').getall()) or ''.join(
            resp.css('td:contains("Téléphone") + td::text').getall()) or sel.xpath(
            '//tr[td[1][contains(text(), "Telephone")]]/td[2]/text()').get('')
        email = resp.css('td:contains("@") ::text').get('')

        item = OrderedDict()
        item['Full Name'] = self.normalize_text(response.meta.get('row', {}).get('Full Name', ''))
        item['Email Address'] = email
        item['Contact'] = telephone
        item['Address'] = self.normalize_text(business_address)
        item['URL'] = response.meta.get('row', {}).get('URL', '')

        if item['Email Address'] or item['Contact']:
            self.item_scraped += 1
            print('Brokers Are Scraped :', self.item_scraped)
            yield item
        else:
            a = 1


    def read_previous_records(self):
        records = {}
        try:
            filenames = glob.glob('outputs/French*.xlsx')  # more flexible
            if not filenames:
                print("[INFO] No previous Excel files found.")
                return {}

            for filepath in filenames:
                print(f"[INFO] Reading: {filepath}")
                workbook = load_workbook(filepath)
                sheet = workbook.active
                headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
                print(f"[DEBUG] Headers: {headers}")

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    url = row_data.get('URL') or row_data.get('url')
                    if url and url not in records:
                        records[url] = row_data

        except Exception as e:
            print(f"[ERROR] Failed to read records: {e}")

        return records

    def normalize_text(self, text):
        if not isinstance(text, str):
            return text

        # Remove HTML entities and unwanted escape characters
        text = html.unescape(text)
        text = text.replace('\xa0', ' ')  # non-breaking space
        text = text.replace('\\"', '"').replace("\\'", "'")
        text = text.replace('\\\\', '\\')
        text = text.replace('\\n', '\n').replace('\\/', '/')

        # Normalize unicode while preserving accents (NFC form keeps é, è, ç etc.)
        text = unicodedata.normalize('NFC', text)

        return text.strip()