import csv
import json, os
from glob import glob
from math import ceil
from datetime import datetime
from urllib.parse import quote
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.reader.excel import load_workbook

from scrapy import Spider, Request, signals
from scrapy.signals import spider_idle


class DealersSpider(Spider):
    name = "Autotrader"
    allowed_domains = ["www.autotrader.co.uk"]
    current_dt = datetime.now().strftime("%d%m%Y%H%M")

    # Scrapy custom settings
    custom_settings = {
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        'DOWNLOAD_TIMEOUT': 70,
        'CONCURRENT_REQUESTS': 4,
    }

    # Headers to mimic browser request
    headers = {
        'accept': '*/*',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'content-type': 'application/json',
        'origin': 'https://www.autotrader.co.uk',
        'pragma': 'no-cache',
        'priority': 'u=1, i',
        'referer': '',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36',
        'x-sauron-app-name': 'sauron-retailer-app',
        'x-sauron-app-version': 'a73aa5bd33',
    }

    def __init__(self):
        super().__init__()

        # Create logs and output directories if not already exist
        os.makedirs('logs', exist_ok=True)
        os.makedirs('output', exist_ok=True)
        self.logs_filepath = f'logs/{self.name} Uk Car dealers logs_{self.current_dt}.txt'
        self.skipped_urls_filepath = f'logs/{self.name} International_skipped_urls.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'[INIT] Script started at {self.script_starting_datetime}')

        # Load previously scraped file For Comparison
        self.previous_file, self.previous_data = self.read_latest_previous_file()

        # Dictionary to hold dealers grouped by postcode
        self.postcode_data = {}

        # List of postcodes (UK cities) to iterate over
        self.post_codes = [('England', 'Bath', 'BA1 1AA'), ('England', 'Birmingham', 'B1 1AA'),
                           ('England', 'Bradford', 'BD1 1AA'), ('England', 'Brighton & Hove', 'BN1 1AA'),
                           ('England', 'Bristol', 'BS1 1AA'), ('England', 'Cambridge', 'CB1 1AA'),
                           ('England', 'Canterbury', 'CT1 1AA'), ('England', 'Carlisle', 'CA1 1AA'),
                           ('England', 'Chelmsford', 'CM1 1AA'), ('England', 'Chester', 'CH1 1AA'),
                           ('England', 'Coventry', 'CV1 1AA'), ('England', 'Derby', 'DE1 1AA'),
                           ('England', 'Doncaster', 'DN1 1AA'), ('England', 'Durham', 'DH1 1AA'),
                           ('England', 'Ely', 'CB7 4AA'), ('England', 'Exeter', 'EX1 1AA'),
                           ('England', 'Gloucester', 'GL1 1AA'), ('England', 'Hereford', 'HR1 1AA'),
                           ('England', 'Kingston-upon-Hull', 'HU1 1AA'),
                           ('England', 'Lancaster', 'LA1 1AA'), ('England', 'Leeds', 'LS1 1AA'),
                           ('England', 'Leicester', 'LE1 1AA'), ('England', 'Lichfield', 'WS13 6AA'),
                           ('England', 'Lincoln', 'LN1 1AA'), ('England', 'Liverpool', 'L1 1AA'),
                           ('England', 'London', 'SW1A 1AA'), ('England', 'Manchester', 'M1 1AA'),
                           ('England', 'Milton Keynes', 'MK1 1AA'), ('England', 'Newcastle-upon-Tyne', 'NE1 1AA'),
                           ('England', 'Norwich', 'NR1 1AA'),
                           ('England', 'Nottingham', 'NG1 1AA'), ('England', 'Oxford', 'OX1 1AA'),
                           ('England', 'Peterborough', 'PE1 1AA'), ('England', 'Plymouth', 'PL1 1AA'),
                           ('England', 'Portsmouth', 'PO1 1AA'), ('England', 'Preston', 'PR1 1AA'),
                           ('England', 'Salford', 'M5 4AA'), ('England', 'Salisbury', 'SP1 1AA'),
                           ('England', 'Sheffield', 'S1 1AA'), ('England', 'Southampton', 'SO14 1AA'),
                           ('England', 'Southend-on-Sea', 'SS1 1AA'),
                           ('England', 'St Albans', 'AL1 1AA'), ('England', 'Stoke-on-Trent', 'ST1 1AA'),
                           ('England', 'Sunderland', 'SR1 1AA'), ('England', 'Truro', 'TR1 1AA'),
                           ('England', 'Wakefield', 'WF1 1AA'), ('England', 'Wells', 'BA5 2AA'),
                           ('England', 'Westminster', 'SW1A 1AA'), ('England', 'Winchester', 'SO23 8AA'),
                           ('England', 'Wolverhampton', 'WV1 1AA'), ('England', 'Worcester', 'WR1 1AA'),
                           ('England', 'York', 'YO1 1AA'), ('Northern Ireland', 'Armagh', 'BT61 7AA'),
                           ('Northern Ireland', 'Bangor', 'BT20 4AA'), ('Northern Ireland', 'Belfast', 'BT1 1AA'),
                           ('Northern Ireland', 'Lisburn', 'BT27 4AA'),
                           ('Northern Ireland', 'Londonderry', 'BT48 7AA'), ('Northern Ireland', 'Newry', 'BT34 1AA'),
                           ('Scotland', 'Aberdeen', 'AB10 1AA'), ('Scotland', 'Dundee', 'DD1 1AA'),
                           ('Scotland', 'Dunfermline', 'KY12 7AA'), ('Scotland', 'Edinburgh', 'EH1 1AA'),
                           ('Scotland', 'Glasgow', 'G1 1AA'), ('Scotland', 'Inverness', 'IV1 1AA'),
                           ('Scotland', 'Perth', 'PH1 1AA'),
                           ('Scotland', 'Stirling', 'FK8 1AA'), ('Wales', 'Bangor', 'LL57 1AA'),
                           ('Wales', 'Cardiff', 'CF10 1AA'), ('Wales', 'Newport', 'NP10 1AA'),
                           ('Wales', 'St Asaph', 'LL17 0AA'),
                           ('Wales', 'St Davids', 'SA62 6AA'), ('Wales', 'Swansea', 'SA1 1AA'),
                           ('Wales', 'Wrexham', 'LL13 8AA')]

        self.len_post_codes = len(self.post_codes)

        # Progress counters
        self.post_code_found = 0
        self.post_code_scrape = 0
        self.all_postcode_found = 0
        self.all_postcode_scrape = 0
        self.empty_postcodes_count = 0
        self.empty_postcodes_list = []

        self.current_scrapeing_postcode = ''
        self.seen_urls = {row['postcode']: row for row in self.read_csv_file()}

    def parse_postcode(self, response):
        # Extract postcode info from meta
        country = response.meta.get('postcode_info', {}).get('country', '')
        city = response.meta.get('postcode_info', {}).get('city', '')
        postcode = response.meta.get('postcode_info', {}).get('postcode', '')
        make = response.meta.get('make', '')
        model = response.meta.get('model', '')
        log_statement = f'Country: {country} | City: {city} | Postcode: {postcode} |'

        try:
            data_dict = response.json()

            # Handle case where 'data' is None
            if not data_dict.get('data'):
                self.write_logs(f"[INFO] {log_statement} No records found")
                self.empty_postcodes_count += 1
                self.empty_postcodes_list.append(postcode)
                return

            dealer_dict = data_dict.get('data', {}).get('search', {}).get('dealers', {}).get('dealerList', {})

        except json.JSONDecodeError as e:
            self.write_logs(f"[ERROR]{log_statement} Function parse_postcode JSON error → {str(e)}")
            return
        except Exception as e:
            self.write_logs(f"[ERROR]{log_statement} Unexpected error in parse_postcode → {str(e)}")
            return

        total_records = dealer_dict.get('totalElements', 0)
        total_pages = ceil(int(total_records) / 12)

        if not response.meta.get('pagination', ''):
            # Log dealer count found per postcode
            self.post_code_found += total_records
            self.all_postcode_found += total_records
            page_word = "page" if total_pages == 1 else "pages"
            self.write_logs(
                f"[DEALERS FOUND] {log_statement} → Total Dealers: {total_records} across {total_pages} {page_word}")

        makes = ['ABARTH', 'AC', 'AK', 'ALFA ROMEO', 'ALPINE', 'ARIEL', 'ASTON MARTIN', 'AUDI', 'AUSTIN', 'BAC',
                 'BENTLEY', 'BMW', 'BUGATTI', 'BYD', 'CADILLAC', 'CATERHAM', 'CHEVROLET', 'CHRYSLER', 'CITROEN',
                 'CUPRA',
                 'DACIA', 'DAIHATSU', 'DAIMLER', 'DAX', 'DODGE', 'DS AUTOMOBILES', 'FERRARI', 'FIAT', 'FISKER', 'FORD',
                 'GENESIS', 'GMC', 'GREAT WALL', 'GWM', 'GWM ORA', 'HONDA', 'HUMMER', 'HYUNDAI', 'INEOS', 'INFINITI',
                 'ISO', 'ISUZU', 'IVECO', 'JAECOO', 'JAGUAR', 'JEEP', 'JENSEN', 'KGM', 'KIA', 'KOENIGSEGG', 'LADA',
                 'LAMBORGHINI', 'LANCIA', 'LAND ROVER', 'LEAPMOTOR', 'LEVC', 'LEXUS', 'LEYLAND', 'LINCOLN',
                 'LONDON TAXIS INTERNATIONAL', 'LOTUS', 'MASERATI', 'MAXUS', 'MAYBACH', 'MAZDA', 'MCLAREN',
                 'MERCEDES-BENZ', 'MG', 'MICRO', 'MINI', 'MITSUBISHI', 'MOKE', 'MORGAN', 'MORRIS', 'NISSAN', 'NOBLE',
                 'OMODA', 'PERODUA', 'PEUGEOT', 'PILGRIM', 'PLYMOUTH', 'POLESTAR', 'PONTIAC', 'PORSCHE', 'PROTON',
                 'RADICAL', 'RBW', 'RENAULT', 'ROBIN HOOD', 'ROLLS-ROYCE', 'ROVER', 'SAAB', 'SEAT', 'SHELBY', 'SKODA',
                 'SKYWELL', 'SMART', 'SSANGYONG', 'SUBARU', 'SUZUKI', 'TESLA', 'TOYOTA', 'TRIUMPH', 'TVR', 'VAUXHALL',
                 'VOLKSWAGEN', 'VOLVO', 'WESTFIELD', 'XPENG']

        models = ['124 SPIDER', '500', '500C', '500E', '500E C', '595', '595C', '600E', '695', '695C', 'GRANDE PUNTO']

        # if total_pages >= 6 and not response.meta.get('make', '') and not response.meta.get('pagination', ''):
        if total_pages >= 16 and not response.meta.get('pagination', ''):
            for make in makes:
                response.meta['make'] = make
                response.meta['pagination'] = True
                json_data = self.get_params_json(postcode, page=1, make=make)
                self.headers[
                    'referer'] = f'https://www.autotrader.co.uk/cars/dealers/search?advertising-location=at_cars&dealerName=&forSale=on&make={quote(make)}&model=&page=1&postcode={quote(postcode)}&radius=100&sort=with-retailer-reviews&toOrder=on'
                yield Request(url='https://www.autotrader.co.uk/at-graphql?opname=RetailerSearchQuery&opname=RetailerSearchWithFacetsQuery',
                              method='POST', dont_filter=True, body=json.dumps(json_data),
                              meta=response.meta, callback=self.parse_postcode, headers=self.headers)

        else:
            # Loop over each page to get full dealer listings
            for page_no in range(1, total_pages + 1):
                if page_no ==17:
                    return

                json_data = self.get_params_json(postcode, page=page_no, make=make if make else None, model=model if model else None)
                response.meta['page_no'] = page_no

                url = f'https://www.autotrader.co.uk/cars/dealers/search?advertising-location=at_cars&dealerName=&forSale=on&make={make if make else ""}&model={model if model else ""}&page={page_no}&postcode={quote(postcode)}&radius=100&sort=with-retailer-reviews&toOrder=on'
                self.headers['referer'] = url
                yield Request(url='https://www.autotrader.co.uk/at-graphql?opname=RetailerSearchQuery', method='POST',
                              dont_filter=True, headers=self.headers, body=json.dumps(json_data),
                              meta=response.meta, callback=self.parse_postcode_listings)

    def parse_postcode_listings(self, response):
        country = response.meta.get('postcode_info', {}).get('country', '')
        city = response.meta.get('postcode_info', {}).get('city', '')
        postcode = response.meta.get('postcode_info', {}).get('postcode', '')
        make = response.meta.get('make', '')
        model = response.meta.get('model', '')
        page_no = response.meta.get('page_no', 0)
        log_statement = f'Country: {country} | City: {city} | Postcode: {postcode} | Make:{make} | Model: {model} | Page No: {page_no} '

        # Initialize postcode entry if not present
        if postcode not in self.postcode_data:
            self.postcode_data[postcode] = []

        try:
            data_dict = response.json()
            dealer_dict = data_dict.get('data', {}).get('search', {}).get('dealers', {}).get('dealerList', {}).get(
                'dealers', [])
        except json.JSONDecodeError as e:
            self.write_logs(f"[ERROR]{log_statement} Dealer Listing Page json response failed → {str(e)}")
            return

        dealer_list = []
        prev_dealer_urls = [url.get('Direct AutoTrader Dealer Page URL') for url in self.postcode_data[postcode]]

        for dealer in dealer_dict:
            try:
                item = OrderedDict()

                # Extract dealer data
                name = dealer.get('name', '')
                d_id = dealer.get('dealerId', '')
                media = dealer.get('media', {})
                website_info = media.get('dealerWebsite')
                formatted_name = name.lower().replace(' ', '-')
                url = f'https://www.autotrader.co.uk/dealers/{formatted_name}-{d_id}?channel=cars'
                star_rating = dealer.get('reviews', {}).get('overallReviewRating', '')
                reviews= dealer.get('reviews', {}).get('numberOfReviews', '')
                no_of_cars = dealer.get('stockLevels', {}).get('atStockCounts', {}).get('stock_car', 0)
                item['Dealer Name'] = name
                item['Address'] = self.get_address(dealer)
                item['Website'] = website_info.get('href') if isinstance(website_info, dict) else ''
                item['Phone Number'] = dealer.get('media', {}).get('phoneNumber1', '')
                item['Star Rating'] = star_rating
                item['Number of Reviews'] = reviews
                item['Number of Cars in Stock'] = str(no_of_cars) if no_of_cars !=0 else ''
                item['Direct AutoTrader Dealer Page URL'] = url
                if url in prev_dealer_urls:  #Avoid Duplications
                    continue

                # Progress tracking
                self.post_code_scrape += 1
                self.all_postcode_scrape += 1
                print(f'{log_statement} Dealers Scrape {self.post_code_scrape}/ {self.post_code_found}')

                dealer_list.append(item)
            except Exception as e:
                self.write_logs(f"[ERROR]{log_statement} Dealer parsing failed → {str(e)}")

        # Save collected dealers to the postcode-specific list
        if dealer_list:
            self.postcode_data[postcode].extend(dealer_list)
        else:
            a = 1

    def get_params_json(self, postcode, page, make=None, model=None):
        limit = 12

        if model or make:
            model = model if model else None
            make = make if make else None
            json_data = [
                {
                    'operationName': 'RetailerSearchQuery',
                    'variables': {
                        'dealerQuery': {
                            'retailerTypes': [
                                'used_car',
                                'new_car',
                            ],
                            'name': '',
                            'postcode': str(postcode),
                            'distance': 100,
                            'make': [make,],
                            'model': [model,] if model else None,
                        },
                        'sort': 'WITH_RETAILER_REVIEWS',
                        # 'limit': 12,
                        'limit': limit,
                        'page': page,
                    },
                    'query': 'query RetailerSearchQuery($dealerQuery: DealerQuery!, $sort: DealerSort, $cursor: String, $page: Int, $limit: Int) {\n  search {\n    dealers(dealerQuery: $dealerQuery, sort: $sort) {\n      dealerList(cursor: $cursor, page: $page, limit: $limit) {\n        size\n        totalElements\n        previousCursor\n        nextCursor\n        dealers {\n          distance\n          dealerId\n          name\n          location {\n            addressOne\n            addressTwo\n            town\n            county\n            region\n            country\n            postcode\n            latLong\n            __typename\n          }\n          marketing {\n            profile\n            brandingBanner {\n              href\n              __typename\n            }\n            strapline\n            __typename\n          }\n          servicesOffered {\n            hasDealerProfilePage\n            __typename\n          }\n          media {\n            phoneNumber1\n            dealerWebsite {\n              href\n              __typename\n            }\n            __typename\n          }\n          reviews {\n            overallReviewRating\n            numberOfReviews\n            __typename\n          }\n          stockLevels {\n            matchingStockCount\n            atStockCounts {\n              car\n              bike\n              van\n              motorhome\n              caravan\n              truck\n              plant\n              agricultural\n              stock_car\n              stock_bike\n              stock_van\n              stock_caravan\n              stock_motorhome\n              __typename\n            }\n            __typename\n          }\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}',
                },
                {
                    'operationName': 'RetailerSearchWithFacetsQuery',
                    'variables': {
                        'dealerQuery': {
                            'retailerTypes': [
                                'used_car',
                                'new_car',
                            ],
                            'name': '',
                            'postcode': str(postcode),
                            'distance': 100,
                            'make': [str(make),],
                            'model': [model,] if model else None,
                        },
                        'sort': 'WITH_RETAILER_REVIEWS',
                        'facets': [
                            'make',
                            'model',
                        ],
                    },
                    'query': 'query RetailerSearchWithFacetsQuery($dealerQuery: DealerQuery!, $sort: DealerSort, $cursor: String, $facets: [SearchFacetName!]!) {\n  search {\n    dealers(dealerQuery: $dealerQuery, sort: $sort) {\n      facets(facets: $facets) {\n        name\n        values {\n          name\n          value\n          __typename\n        }\n        __typename\n      }\n      dealerList(cursor: $cursor) {\n        totalElements\n        previousCursor\n        nextCursor\n        dealers {\n          distance\n          dealerId\n          name\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}',
                },
            ]

        else:
            # Return JSON body for POST request for dealer search
            json_data = [
                {
                    'operationName': 'RetailerSearchQuery',
                    'variables': {
                        'dealerQuery': {
                            'retailerTypes': [
                                'used_car',
                                'new_car',
                            ],
                            'name': '',
                            'postcode': postcode,
                            'distance': 100,
                            'make': None,
                            'model': None,
                        },
                        'sort': 'WITH_RETAILER_REVIEWS',
                        # 'limit': 12,
                        'limit': limit,
                        'page': page,
                    },
                    'query': 'query RetailerSearchQuery($dealerQuery: DealerQuery!, $sort: DealerSort, $cursor: String, $page: Int, $limit: Int) {\n  search {\n    dealers(dealerQuery: $dealerQuery, sort: $sort) {\n      dealerList(cursor: $cursor, page: $page, limit: $limit) {\n        size\n        totalElements\n        previousCursor\n        nextCursor\n        dealers {\n          distance\n          dealerId\n          name\n          location {\n            addressOne\n            addressTwo\n            town\n            county\n            region\n            country\n            postcode\n            latLong\n            __typename\n          }\n          marketing {\n            profile\n            brandingBanner {\n              href\n              __typename\n            }\n            strapline\n            __typename\n          }\n          servicesOffered {\n            hasDealerProfilePage\n            __typename\n          }\n          media {\n            phoneNumber1\n            dealerWebsite {\n              href\n              __typename\n            }\n            __typename\n          }\n          reviews {\n            overallReviewRating\n            numberOfReviews\n            __typename\n          }\n          stockLevels {\n            matchingStockCount\n            atStockCounts {\n              car\n              bike\n              van\n              motorhome\n              caravan\n              truck\n              plant\n              agricultural\n              stock_car\n              stock_bike\n              stock_van\n              stock_caravan\n              stock_motorhome\n              __typename\n            }\n            __typename\n          }\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}',
                },
            ]

        json_data = json_data[0]
        return json_data

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def get_address(self, dealer):
        location = dealer.get('location', {})
        parts = [
            location.get('addressOne', ''),
            location.get('town', ''),
            location.get('county', ''),
            location.get('postcode', '')
        ]

        # Filter out empty parts to avoid blank lines
        address = '\n'.join(filter(None, parts))
        return address

    def read_latest_previous_file(self):
        files = glob('output/*.xlsx')
        files = [f for f in files if self.name in f]
        if not files:
            return None, {}

        files.sort(key=lambda x: os.path.getmtime(x), reverse=True)

        # Get previous file (skip current one if re-running same day)
        previous_file = files[1] if len(files) > 1 else files[0]

        data = {}
        try:
            wb = load_workbook(previous_file, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(cell).strip() if cell else '' for cell in rows[0]]
                records = []
                for row in rows[1:]:
                    record = {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers))}
                    records.append(record)
                data[sheet] = records
        except Exception as e:
            print(f"Error reading {previous_file}: {e}")

        return previous_file, data

    def export_to_excel(self):
        filename = f'output/{self.name}_UK_Car_Dealers_{self.current_dt}.xlsx'
        # filename = f'output/Autotrader_Car_Dealers_080520252240.xlsx'
        headers = ['Dealer Name', 'Address', 'Website', 'Phone Number','Star Rating', 'Number of Reviews', 'Number of Cars in Stock', 'Direct AutoTrader Dealer Page URL']
        postcode = ''

        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # New
        purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')  # Removed

        # Check if file exists, load it; otherwise, create new workbook
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet only if creating new file

        for postcode, records in self.postcode_data.items():
            try:
                sheet_name = postcode[:31]  # Excel sheet name max length = 31

                if sheet_name in wb.sheetnames:
                    self.write_logs(
                        f"[SKIP] Sheet for postcode {postcode} already exists. Skipping to avoid duplication.")
                    continue  # Skip if sheet already exists

                ws = wb.create_sheet(title=sheet_name)
                ws.append(headers)

                current_records = []
                for item in records:
                    row = tuple(item.get(header, '') for header in headers)
                    current_records.append(row)

                prev_records = self.previous_data.get(postcode, set())

                # Create dict for previous data by URL for easy lookup
                previous_data_by_url = {
                    dealer['Direct AutoTrader Dealer Page URL']: dealer
                    for dealer in prev_records
                    if dealer.get('Direct AutoTrader Dealer Page URL', '')
                }

                # Create dict for current data by URL
                current_data_by_url = {
                    dealer['Direct AutoTrader Dealer Page URL']: dealer
                    for dealer in records
                    if dealer.get('Direct AutoTrader Dealer Page URL')
                }

                # For logging counts
                new_entries = 0
                existing_entries = 0
                removed_entries = 0

                for item in records:  #New Entries: Highlight in yellow
                    row = tuple(item.get(header, '') for header in headers)
                    url = item.get('Direct AutoTrader Dealer Page URL', '')
                    ws.append(row)
                    dealer_name_cell = ws[ws.max_row][0]

                    if url not in previous_data_by_url:
                        dealer_name_cell.fill = yellow_fill
                        new_entries += 1
                    else:
                        existing_entries += 1

                # 2. Add removed records (those in previous but not in current)
                removed_urls = set(previous_data_by_url.keys()) - set(current_data_by_url.keys())
                for url in removed_urls:
                    removed_dealer = previous_data_by_url[url]
                    row = tuple(removed_dealer.get(header, '') for header in headers)
                    ws.append(row)
                    dealer_name_cell = ws[ws.max_row][0]
                    dealer_name_cell.fill = purple_fill
                    removed_entries += 1

                self.write_logs(
                    f"[EXPORT] {postcode} — Total: {len(current_records) + removed_entries}, "
                    f"New: {new_entries}, Existing: {existing_entries}, Removed: {removed_entries}"
                )

            except Exception as e:
                self.write_logs(f"[ERROR] Failed to write records for {postcode}: {str(e)}")

        wb.save(filename)
        self.write_logs(f"[EXPORT COMPLETE] Postcode: {postcode} | All dealer records saved to {filename}")
        self.postcode_data = {}

    def write_item_into_csv_file(self, item):
        # to ensure that all  directories are exists
        output_filename = 'output/seen postcode.csv'
        os.makedirs(os.path.dirname(output_filename), exist_ok=True)
        fieldnames = item.keys()

        with open(output_filename, mode='a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            if csvfile.tell() == 0:
                writer.writeheader()

            writer.writerow(item)

    def read_csv_file(self):
        try:
            with open('output/seen postcode.csv', mode='r', encoding='utf-8') as csv_file:
                return list(csv.DictReader(csv_file))
        except:
            return []

    def close(self, reason):
        # self.export_to_excel()
        self.write_logs("[FINISHED] Data exported for all postcodes.")
        """Shutdown logic when the spider closes."""

        # Logging the summary
        self.write_logs("=" * 60)
        self.write_logs(f"\n--- Scraping Summary ---")
        self.write_logs(f"📌 Total Postcodes Processed: {self.len_post_codes}")
        self.write_logs(f"🏷️ Postcodes with Dealer Records: {self.len_post_codes - self.empty_postcodes_count}")
        self.write_logs(f"🚫 Postcodes with No Dealers Found: {self.empty_postcodes_count}")
        self.write_logs(f"👥 Total Dealer found: {self.all_postcode_found}")
        self.write_logs(f"👥 Total Dealer Records Collected: {self.all_postcode_scrape}")

        if self.empty_postcodes_list:
            self.write_logs(f"❗ Empty Postcodes: {', '.join(self.empty_postcodes_list)}")

        # Log script execution times
        self.write_logs(f"\n--- Script Execution Times ---")
        self.write_logs(f"🛑 Spider '{self.name}' has finished running.")
        self.write_logs(f"📅 Start Time: {self.script_starting_datetime}")
        self.write_logs(f"🕓 End Time: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        self.write_logs(f"🔚 Closing Reason: {reason}")
        self.write_logs("=" * 60)

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(DealersSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        if self.post_code_scrape >= 1:
            self.export_to_excel()
            self.write_item_into_csv_file(item={'postcode': self.current_scrapeing_postcode})

        # if self.post_codes:
        while self.post_codes:
            self.write_logs(f"\n\n{len(self.post_codes)}/{self.len_post_codes} Post Codes left to Scrape\n")

            self.post_code_found = 0
            self.post_code_scrape = 0

            postcode_tuple = self.post_codes.pop()
            country = postcode_tuple[0]
            city = postcode_tuple[1]
            postcode = postcode_tuple[2]
            self.current_scrapeing_postcode = postcode

            postcode_info = {
                'country': country,
                'city': city,
                'postcode': postcode
            }

            if postcode in self.seen_urls:
                print('Already Scraped')
                continue

            self.write_logs(f'Country :{country} | City :{city} | Postcode :{postcode} started scrape')
            url = f'https://www.autotrader.co.uk/cars/dealers/search?advertising-location=at_cars&dealerName=&forSale=on&make=&model=&page=1&postcode={quote(postcode)}&radius=100&sort=with-retailer-reviews&toOrder=on'
            self.headers['referer'] = url
            json_data = self.get_params_json(postcode, page=1)

            self.crawler.engine.crawl(Request(url='https://www.autotrader.co.uk/at-graphql?opname=RetailerSearchQuery',
                                              method='POST', dont_filter=True, headers=self.headers,
                                              callback=self.parse_postcode,
                                              body=json.dumps(json_data), meta={'postcode_info': postcode_info}))

            break  # Exit the loop once a valid crawl has been scheduled


