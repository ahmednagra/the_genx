import openpyxl
import glob
import os

# This will create a new workbook to store removed rows
removed_wb = openpyxl.Workbook()
removed_ws = removed_wb.active
removed_ws.title = "Removed Dealers"
removed_written_header = False


# Dealers and brands to remove
remove_dealers = ['Sytner Group', 'Arnold Clark Automobiles', 'Lookers', 'Vertu Motors', 'Pendragon', 'Evans Halshaw',
                  'Stratstone', 'Citygate', 'Marshall Motor Group', 'Group 1 Automotive', 'Barons Group', 'Think Ford',
                  'TrustFord', 'Jardine Motors Group', 'JCT600', 'Inchcape', 'Listers', 'PSA Retail UK', 'Stoneacre Motor Group',
                  'Motus UK (Pentagon Group)', 'Greenhous Group', 'Hendy Group', 'Rybrook Holdings', 'Hedin Automotive UK',
                  'John Clark Motor Group', 'Park\'s Motor Group', 'Swansway Garages', 'Eastern Western Motor Group',
                  'Johnsons Cars', 'Dick Lovett Group', 'Lloyd Motors', 'Steven Eagell Group', 'RRG Group', 'Norton Way Motors',
                  'Holdcroft Motor Group', 'Perrys Group', 'Harwoods Group', 'Peter Vardy', 'Snows Motor Group',
                  'Cambria Automobiles', 'Sinclair Group', 'AWR Holdings UK', 'Williams Motor Company', 'HR Owen', 'Allen Ford',
                  'Glyn Hopkin', 'Halliwell Jones', 'Renault Retail Group', 'Marsh Wall', 'Vindis Group', 'Citygate Automotive',
                  'LSH Auto UK', 'TC Harrison Group', 'Hatfields', 'Porsche Retail Group', 'Yeomans', 'Mon Motors',
                  'Jemca Car Group', 'Richmond Motor Group', 'VT Holdings', 'Bowker Motor Group', 'Drive Motor Retail',
                  'Peoples Ford', 'Sandicliffe Motor Group', 'Eden Motor Group', 'Barretts of Canterbury', 'CEM Day Motor Group',
                  'BMW Park Lane', 'Donnelly Group', 'Vospers Motor House', 'Caffyns', 'Brindley Garages Group', 'DM Keith',
                  'Thurlow Nunn', 'William Morgan Group', 'City West Country', 'Endeavour Automotive', 'Chorley Group',
                  'Toomey Motor Group', 'Hartwell PLC', 'JCB Group', 'Sandown Motors', 'Vantage Motor Group', 'Vines Group',
                  'Cotswold Motor Group', 'Partridge of Hampshire', 'Riverside Motor Group', 'Fish Brothers', 'SG Petch',
                  'Gates Ford', 'Foray Motor Group', 'John Grose Group', 'Parkway Motor Group', 'Arbury Motor Group',
                  'Heritage Automotive', 'Ancaster Group', 'Marriott Motor Group', 'Howards Motor Group', 'Ocean Automotive',
                  'Burrows Motor Company', 'Waylands Automotive', 'Cars2', 'Central Garage Uppingham', 'RJTK Investments',
                  'Read Motor Group', 'Drift Bridge Group', 'Jennings Motor Group', 'Spire Automotive', 'Beadles Group',
                  'Robinsons Motor Group', 'Lancaster Motor Group', 'Appleyard Blight', 'Polar Motor Group', 'Milcars BMW',
                  'Toyota & Lexus Retailers', 'Scotthall', 'Minories', 'Abridge Garage', 'Clover Leaf Cars', 'Elms BMW',
                  'Essex Audi', 'Mercedes-Benz Retail Group', 'Smart Retail Group', 'Abarth', 'Alfa Romeo', 'Audi', 'Bentley',
                  'BMW', 'BYD', 'Citroën', 'Cupra', 'Dacia', 'DS Automobiles', 'Ferrari', 'Fiat', 'Ford', 'Genesis', 'GWM ORA',
                  'Honda', 'Hyundai', 'Infiniti (legacy, few still supported)', 'Isuzu', 'Jaguar', 'Jeep', 'Kia', 'Lamborghini',
                  'Land Rover', 'Lexus', 'Lotus', 'Maserati', 'Mazda', 'McLaren', 'Mercedes-Benz', 'MG', 'Mini',
                  'Mitsubishi (legacy support)', 'Nissan', 'Peugeot', 'Polestar', 'Porsche', 'Renault', 'Rolls-Royce',
                  'Saab', 'SEAT', 'Škoda', 'Smart', 'SsangYong', 'Subaru', 'Suzuki', 'Tesla', 'Toyota', 'Vauxhall',
                  'Volkswagen', 'Volvo']


def clean_excel_file(file_path, remove_dealers):
    global removed_written_header

    wb = openpyxl.load_workbook(file_path)
    dealer_field = "Dealer Name"
    remove_dealers_set = {d.strip().lower() for d in remove_dealers}

    for sheet in wb.worksheets:
        print(f"Processing sheet: {sheet.title}")

        # Get header and dealer column index
        header = [cell.value for cell in sheet[1]]
        if dealer_field not in header:
            print(f"'{dealer_field}' column not found in sheet: {sheet.title}")
            continue

        dealer_idx = header.index(dealer_field)
        rows_to_delete = []

        # Write header once to the removed dealers file
        if not removed_written_header:
            removed_ws.append(header)
            removed_written_header = True

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            dealer_name = row[dealer_idx].value.strip().lower()
            if dealer_name in remove_dealers_set:
                removed_ws.append([cell.value for cell in row])
                rows_to_delete.append(i)

        # Delete rows in reverse to avoid index shifting
        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)

    wb.save(file_path)
    print(f"Cleaned and saved: {file_path}\n")


def main():
    excel_file = glob.glob("output/Autotrader_UK_Car_Dealers_150520251848.xlsx")[0]
    if not excel_file:
        print("No Excel files found in 'output/' folder.")
        return

    print(f"\nOpening file: {os.path.basename(excel_file)}")
    clean_excel_file(excel_file, remove_dealers)

    # Save the removed dealers file
    removed_wb.save("AutotraderUK_RemovedDealers.xlsx")
    print("Saved all removed dealer rows to 'AutotraderUK_RemovedDealers.xlsx'")


if __name__ == "__main__":
    main()