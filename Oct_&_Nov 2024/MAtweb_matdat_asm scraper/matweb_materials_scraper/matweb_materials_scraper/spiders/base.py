import json
import os
import csv
import random
import ipaddress
from datetime import datetime
from scrapy import Spider

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class BaseSpider(Spider):
    name = 'base'
    current_dt = datetime.now().strftime("%d%m%Y%H%M")

    custom_settings = {
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        # 'CONCURRENT_REQUESTS': 2,
    }

    fields = [
        # headers for Matweb.com
        'Name', 'Categories', 'Key Words', 'Vendors', 'Color', 'Crystal Structure',
        'Physical Properties', 'Chemical Properties', 'Mechanical Properties',
        'Electrical Properties', 'Thermal Properties', 'Optical Properties',
        'Component Elements Properties', 'Descriptive Properties', 'Processing Properties',
        'Chemical Resistance Properties', 'URL',
        # Headers for makeitfrom.com
        'Material', 'url', 'Description', 'Alloy Composition', 'Base Metal Price (% relative)',
        'Brinell Hardness', 'Compressive (crushing) Strength (MPa)', 'Curie Temperature (°C)',
        'Density (g/cm³)', 'Dielectric Constant (relative Permittivity) At 1Hz',
        'Dielectric Constant (relative Permittivity) At 1Mhz',
        'Dielectric Strength (breakdown Potential) (kV/mm)',
        "Elastic (Young's, Tensile) Modulus (GPa)", 'Electrical Conductivity: Equal Volume (% IACS)',
        'Electrical Conductivity: Equal Weight (specific) (% IACS)', 'Electrical Dissipation At 1Hz',
        'Electrical Dissipation At 1Mhz', 'Electrical Resistivity Order Of Magnitude (10)',
        'Elongation At Break (%)',
        'Embodied Carbon (kg CO₂)', 'Embodied Energy (MJ/kg)', 'Embodied Water (L/kg)',
        'Fatigue Strength (MPa)',
        'Flexural Modulus (GPa)', 'Flexural Strength (MPa)', 'Follow-up Questions',
        'Fracture Toughness (MPa·m)',
        'Further Reading', 'Glass Transition Temperature (°C)',
        'Heat Deflection Temperature At 1.82Mpa (264Psi) (°C)',
        'Impact Strength: Notched Izod (J/m)', 'Impact Strength: V-notched Charpy (J)', 'Knoop Hardness',
        'Latent Heat Of Fusion (J/g)', 'Light Transmission Range (µm)', 'Limiting Oxygen Index (loi) (%)',
        'Maximum Temperature: Autoignition (°C)', 'Maximum Temperature: Corrosion (°C)',
        'Maximum Temperature: Decomposition (°C)',
        'Maximum Temperature: Mechanical (°C)', 'Maximum Thermal Shock (°C)',
        'Melting Completion (liquidus) (°C)',
        'Melting Onset (solidus) (°C)', "Poisson's Ratio", 'Pren (pitting Resistance)',
        'Reduction In Area (%)',
        'Refractive Index', 'Resilience: Ultimate (unit Rupture Work) (MJ/m³)',
        'Resilience: Unit (modulus Of Resilience) (kJ/m³)',
        'Rockwell B Hardness', 'Rockwell C Hardness', 'Rockwell M Hardness', 'Shear Modulus (GPa)',
        'Shear Strength (MPa)',
        "Solidification (pattern Maker's) Shrinkage (%)", 'Specific Heat Capacity (J/kg·K)',
        'Stiffness To Weight: Axial (points)',
        'Stiffness To Weight: Bending (points)', 'Strength To Weight: Axial (points)',
        'Strength To Weight: Bending (points)',
        'Tensile Strength: Ultimate (uts) (MPa)', 'Tensile Strength: Yield (proof) (MPa)',
        'Thermal Conductivity (W/m·K)',
        'Thermal Diffusivity (mm²/s)', 'Thermal Expansion (µm/m·K)', 'Thermal Shock Resistance (points)',
        'Vicat Softening Temperature (°C)', 'Water Absorption After 24 Hours (%)',
        'Water Absorption At Saturation (%)',
        # MADDAT Headers
        'Material ID (MAT_ID)', 'Material Designation', 'Manufacturer/Supplier', 'Chemical Composition',
        'Semifinished Material Information', 'Heat Treatment', 'Microstructure', 'Hardness', 'Testing Conditions (Axial Loading)',
        "Young's Modulus (N/mm²)", "Poisson's Ratio", 'Yield Strength (N/mm²)', 'Ultimate Tensile Strength (N/mm²)',
        'Elongation (A5, %)', 'Reduction of Area (RA, %)', 'True Fracture Stress (N/mm²)', 'True Fracture Strain',
        'Monotonic Stress-Strain Curves (Ramberg-Osgood Model)', 'Cyclic/Fatigue Properties (Axial Loading, Fully Reversed)',
        'Testing Temperature (°C)', 'Testing Medium', 'Loading Type', 'Loading Control', 'Specimen', 'Loading Ratio',
        'Additional Remarks', 'Fatigue Properties', 'Fatigue Strain-Life Parameters (Coffin-Manson-Basquin Model)',
        'Collecting Cyclic Stress-Strain Plot Data', 'Collecting Strain-Life Fatigue Plot Data',
        # ASM Internation
        'Type', 'Material Group', 'Title', 'Country/Standard', 'Producer', 'Material group', 'Flexural Strength',
        'Shear Strength', 'Compression Strength', 'Yield Strength', 'Tensile Strength', 'Tensile Strain',
        'Tracker',
        'Cross Referencing', 'Modulus of Elasticity', 'Density', 'Shrinkage', 'Specific Gravity', 'Thermal Conductivity',
        'Glass Tran. Temp'
    ]

    def __init__(self):
        super().__init__()
        # Logs
        os.makedirs('logs', exist_ok=True)
        self.logs_filepath = f'logs/{self.name}_logs {self.current_dt}.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'[INIT] Script started at {self.script_starting_datetime}')

        self.ready_write_records = []
        self.ready_records_counter = 0
        self.all_categories_found_records = 0
        self.all_categories_scraped_records = 0

    def start_requests(self):
        pass

    def parse(self, response, **kwargs):
        pass

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def write_xlsx(self, record, key=False):
        """Write records to the Excel file when the counter reaches 10."""
        output_dir = 'output1'
        os.makedirs(output_dir, exist_ok=True)
        output_file = f'{output_dir}/{self.name} Materials Details.xlsx'

        if record:
            self.ready_write_records.append(record)

        # Only write to the file when we have 10 records
        if len(self.ready_write_records) >= 1000 or key == True:
            file_exists = os.path.exists(output_file)

            if not file_exists:
                wb = Workbook()
                ws = wb.active
                ws.title = "Materials"
                ws.append(self.fields)
            else:
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active

            # Write all 10 records to the file at once
            for record in self.ready_write_records:
                complete_record = [
                    record.get(field, 'N/A') if record.get(field, 'N/A') not in (None, '', ' ') else 'N/A'
                    for field in self.fields
                ]
                ws.append(complete_record)

            wb.save(output_file)

            if 'makeitfrom.com' in record.get('URL', '') :
                for title in [r.get('Material', '') for r in self.ready_write_records]:
                    print(f"Record for '{title}' written to Excel successfully.")
                    self.all_categories_scraped_records += 1

            elif 'tpsx.arc.nasa.gov' in record.get('URL', ''):
                for title in [r.get('Name', '')[:20] for r in self.ready_write_records]:
                    print(f"Record for '{title}' written to Excel successfully.")
                    self.all_categories_scraped_records += 1

            elif 'matdat' in record.get('spider', ''):
                for mat_id in [r.get('Material ID (MAT_ID)', '') for r in self.ready_write_records]:
                    print(f"Record for Material ID:'{mat_id}' written to Excel successfully.")
                    self.all_categories_scraped_records += 1

            self.ready_write_records = []

    def write_json(self, record):
        """Write a single record to the JSON file."""
        output_dir = 'output1'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_file = f'{output_dir}/{self.name} Materials Details {self.current_dt}.json'

        try:
            # Check if the JSON file exists and read its contents
            if os.path.exists(output_file):
                with open(output_file, 'r', encoding='utf-8') as json_file:
                    data = json.load(json_file)  # Load existing records
            else:
                data = []  # Initialize an empty list if the file doesn't exist

            # Prepare the record to be written (same structure as your fields)
            complete_record = {field: record.get(field, 'N/A') for field in self.fields}

            # Append the new record to the data
            data.append(complete_record)

            # Write the updated data to the JSON file
            with open(output_file, 'w', encoding='utf-8') as json_file:
                json.dump(data, json_file, ensure_ascii=False, indent=4)

            self.product_count += 1
            print(f'Items Scraped: {self.product_count}')
            print(f"Record for '{record.get('Name', '')}' written to JSON successfully.")
        except Exception as e:
            self.write_logs(
                f"Title: {record.get('Name', '')} Url:{record.get('URL', '')} Error writing to the JSON file: {e}")

    def close(Spider, reason):
        Spider.write_xlsx(record=False, key=True)

        # Log overall scraping statistics
        Spider.write_logs(f"\n--- Scraping Summary ---")
        Spider.write_logs(f"Total Products Available on Website: {Spider.all_categories_found_records}")
        Spider.write_logs(f"Total Products Successfully Scraped: {Spider.all_categories_scraped_records}")

        # Log script execution times
        Spider.write_logs(f"\n--- Script Execution Times ---")
        Spider.write_logs(f"Script Start Time: {Spider.script_starting_datetime}")
        Spider.write_logs(f"Script End Time: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        Spider.write_logs(f"Reason for Closure: {reason}")

