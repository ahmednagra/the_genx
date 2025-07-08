import os
import json
import glob
from datetime import datetime
from collections import OrderedDict
from urllib.parse import urlparse, parse_qs, urlencode

import requests

from math import ceil

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

from scrapy import Request, Spider, signals, Selector


class MatwebSpider(Spider):
    name = "MatWeb"
    current_dt = datetime.now().strftime("%d%m%Y%H%M")

    custom_settings = {
        'RETRY_TIMES': 3,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        'DOWNLOAD_TIMEOUT': 150,
        'DOWNLOAD_DELAY': 5,  # Wait 5 seconds between requests
        'RANDOMIZE_DOWNLOAD_DELAY': True,  # Add variation to avoid detection

        "ZYTE_API_EXPERIMENTAL_COOKIES_ENABLED": True,
        'DOWNLOAD_HANDLERS': {
            "http": "scrapy_zyte_api.ScrapyZyteAPIDownloadHandler",
            "https": "scrapy_zyte_api.ScrapyZyteAPIDownloadHandler",
        },
        'DOWNLOADER_MIDDLEWARES': {
            "scrapy_zyte_api.ScrapyZyteAPIDownloaderMiddleware": 1000,
            "scrapy_poet.InjectionMiddleware": 543,
        },
        'REQUEST_FINGERPRINTER_CLASS': "scrapy_zyte_api.ScrapyZyteAPIRequestFingerprinter",
        'TWISTED_REACTOR': "twisted.internet.asyncioreactor.AsyncioSelectorReactor",
        'ZYTE_API_KEY': "905eefc5007b4c5b86a06cb416a0061d",
        "ZYTE_API_TRANSPARENT_MODE": True,

    }

    fields = [
        # headers for Matweb.com according new changing every header getting dynamically
        'Name', 'Categories', 'Key Words', 'Vendors', 'URL'
    ]

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'priority': 'u=0, i',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
    }

    def __init__(self):
        super().__init__()
        print('Script is Started')
        self.ready_write_records = []
        self.current_sub_cat_found = 0
        self.current_sub_cat_scraped = 0
        self.all_groups_found_records = 0
        self.all_groups_scraped_records = 0
        self.ready_write_records_counter = 0

        os.makedirs('logs', exist_ok=True)
        os.makedirs('output1', exist_ok=True)
        self.current_scraped_sub_cat = ''
        self.skipped_urls_filepath = f'logs/{self.name}_skipped_urls.txt'
        self.skipped_urls = list(set(self.read_skipped_urls()))

        self.categories = ['carbon', 'ceramic', 'metal', 'polymer']
        self.all_sub_categories = self.read_json_files(filepath='input/matweb_input/sub_categories.json')[0]
        self.all_material_groups = len(self.all_sub_categories)  # for idle function logs
        self.main_url = 'https://www.matweb.com/search/MaterialGroupSearch.aspx'
        self.cert = glob.glob('input/matweb_input/zyte-ca.crt')[0]

        # Logs
        self.output_file_path = f'output1/{self.name} Materials Details.xlsx'
        self.previous_scraped_records = self.read_write_xlsx_file(key='previous_records')
        self.logs_filepath = f'logs/{self.name}_{self.current_dt}.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'[INIT] Script started at {self.script_starting_datetime}')
        self.output_file = f'output1/{self.name} Materials Details.xlsx'

        self.proxy = {
            scheme: "http://905eefc5007b4c5b86a06cb416a0061d:@api.zyte.com:8011" for scheme in ("http", "https")
        }

    def parse_home(self, response):
        req = requests.get(url=self.main_url, proxies=self.proxy, headers=self.headers, verify=self.cert)
        if req.status_code != 200:
            req = requests.get(url=self.main_url, headers=self.headers)

        maintenance_text = 'Matweb is currently undergoing maintenance'
        if maintenance_text in req.text:
            self.write_logs("Maintenance mode detected. Stopping spider.")
            self.crawler.engine.close_spider(self, reason='maintenance')
            return

        sub_category_dict = response.meta.get('sub_category', '')
        parent_category = sub_category_dict.get('parent_category', '')
        sub_cat_name = sub_category_dict.get('GroupName', '')
        info = f'Category: {parent_category} & Sub Category:{sub_cat_name}'

        response = Selector(text=req.text)
        viewstat = response.css('#__VIEWSTATE ::attr("value")').get('')
        viewstat_generator_id = response.css('#__VIEWSTATEGENERATOR ::attr("value")').get('')

        # first get 50 records
        data = self.get_sub_cat_form_data(viewstate=viewstat, view_genrator=viewstat_generator_id,
                                          parent_category=parent_category, sub_category_dict=sub_category_dict,
                                          key='first_page')

        f_page_req = requests.post(url=self.main_url, data=data, proxies=self.proxy, verify=self.cert)
        if f_page_req.status_code != 200:
            f_page_req = requests.post(url=self.main_url, data=data)

        f_page_res = Selector(text=f_page_req.text)
        total_products = f_page_res.css('#ctl00_ContentMain_lblMatlCount ::text').get('') or 0

        self.current_sub_cat_found += int(total_products)
        self.all_groups_found_records += int(total_products)
        print(f'{info} Category: {parent_category}, Current Sub Category:{sub_cat_name} Found record on web: {self.current_sub_cat_found}')

        if int(total_products) >= 50:
            yield from self.parse_pagination(response=f_page_res, total_products=total_products,
                                             sub_category=sub_category_dict)
        else:
            materials_urls = f_page_res.css('#tblResults tr a::attr(href)').getall() or []
            if materials_urls:
                yield from self.parse_detail(urls=materials_urls, key='parse_home', request=f_page_req)

    def parse_pagination(self, response, total_products, sub_category):
        print('Parse parse_pagination')
        sub_category_dict = sub_category
        parent_category = sub_category_dict.get('parent_category', '')
        sub_cat_name = sub_category_dict.get('GroupName', '')

        new_viewstat = response.css('#__VIEWSTATE ::attr("value")').get('')
        new__viewstat_generator_id = response.css('#__VIEWSTATEGENERATOR ::attr("value")').get('')
        data = self.get_sub_cat_form_data(viewstate=new_viewstat, view_genrator=new__viewstat_generator_id,
                                          parent_category=parent_category, sub_category_dict=sub_category_dict,
                                          key='two_hundred', page_no='1')

        all_urls = []
        # this request gets the 200 records from response
        full_page_req = requests.post(url=self.main_url, data=data, proxies=self.proxy, verify=self.cert)
        if full_page_req != 200:
            full_page_req = requests.post(url=self.main_url, data=data)

        full_page_res = Selector(text=full_page_req.text)
        materials_urls = full_page_res.css('#tblResults tr a::attr(href)').getall() or []
        np_viewstat = full_page_res.css('#__VIEWSTATE ::attr("value")').get('')
        if materials_urls:
            yield from self.parse_detail(urls=materials_urls, key='parse_pagination_full_page_res')

        if int(total_products) >= 200:
            self.headers['content-type'] = 'application/x-www-form-urlencoded'
            self.headers['origin'] = 'https://www.matweb.com'
            self.headers['referer'] = 'https://www.matweb.com/search/MaterialGroupSearch.aspx'
            total_prod_remain = int(total_products) - 200 if int(total_products) else 0
            total_pages = ceil(total_prod_remain / 200) if total_prod_remain > 0 else 0

            for page_no in range(1, total_pages + 2):
                if page_no == 6:
                    break

                pagination_data = self.get_sub_cat_form_data(viewstate=np_viewstat,
                                                             view_genrator=new__viewstat_generator_id, parent_category=parent_category,
                                                             # sub_category_dict=sub_category_dict, key='pagination', page_no=page_no)
                                                             sub_category_dict=sub_category_dict, key='two_hundred', page_no=page_no)

                next_page_req = requests.post(url=self.main_url, data=pagination_data, proxies=self.proxy, verify=self.cert)
                if next_page_req != 200:
                    next_page_req = requests.post(url=self.main_url, data=pagination_data)

                next_page_res = Selector(text=next_page_req.text)
                materials_urls = next_page_res.css('#tblResults tr a::attr(href)').getall() or []
                np_viewstat = next_page_res.css('#__VIEWSTATE ::attr("value")').get('')
                if materials_urls:
                    all_urls.extend(materials_urls)

        if all_urls:
            yield from self.parse_detail(urls=list(set(all_urls)), key='parse_pagination_next_page_res',
                                         request=next_page_req)

    def parse_detail(self, urls=None, key=None, request=None):
        print(f'Current Products Called From Function :{key}')
        previous_records_urls = {self.normalize_url(url) for url in self.previous_scraped_records.keys()}

        for mat_url in urls:
            url = f'https://www.matweb.com{mat_url}'
            normalized_url = self.normalize_url(url)
            if normalized_url in previous_records_urls:
                self.current_sub_cat_scraped += 1
                self.all_groups_scraped_records += 1
                print('Material Previously Scraped Skipped..')
                continue

            yield Request(
                url=url,
                callback=self.parse_item,
                meta={
                    'handle_httpstatus_all': True,
                    'redirect_enabled': True,
                    'zyte_api': {
                        "browserHtml": True,  # Enables JavaScript execution
                    }
                },
                # dont_filter=True
            )

    def parse_item(self, response):
        url = response.url

        try:
            item = OrderedDict()
            title = response.css('.tabledataformat.t_ableborder th::text').get('').strip()
            item['Name'] = title
            item['Categories'] = ', '.join(
                response.css('th:contains("Categories:") + td a::text').getall()) or ', '.join(
                response.css('#ctl00_ContentMain_ucDataSheet1_trMatlGroups td a ::text').getall())
            item['Key Words'] = ', '.join(
                response.css('#ctl00_ContentMain_ucDataSheet1_trMatlNotes td ::text').getall())
            item['Vendors'] = self.get_vendor(response)
            item['Physical Properties'] = self.get_physical_properties(item, response, 'Physical Properties')
            # item['Mechanical Properties'] = self.get_properties(item, response, 'Mechanical', 'Mechanical Properties')
            self.get_properties(item, response, 'Mechanical', 'Mechanical Properties')
            # item['Thermal Properties'] = self.get_properties(item, response, 'Thermal', 'Thermal Properties')
            self.get_properties(item, response, 'Thermal', 'Thermal Properties')
            # item['Electrical Properties'] = self.get_properties(item, response, 'Electrical', 'Electrical Properties')
            self.get_properties(item, response, 'Electrical', 'Electrical Properties')
            # item['Chemical Properties'] = self.get_properties(item, response, 'Chemical', 'Chemical Properties')
            self.get_properties(item, response, 'Chemical', 'Chemical Properties')
            # item['Optical Properties'] = self.get_properties(item, response, 'Optical', 'Optical Properties')
            self.get_properties(item, response, 'Optical', 'Optical Properties')
            # item['Processing Properties'] = self.get_properties(item, response, 'Processing', 'Processing Properties')
            self.get_properties(item, response, 'Processing', 'Processing Properties')
            # item['Safety Information'] = self.get_properties(item, response, 'Safety', 'Safety Information')
            self.get_properties(item, response, 'Safety', 'Safety Information')
            # item['Physical Properties'] = self.get_physical_properties(item, response, 'Physical Properties')
            # item['Component Elements Properties'] = self.get_properties(item, response, 'Component Elements','Component Elements Properties')
            self.get_properties(item, response, 'Component Elements', 'Component Elements Properties')
            # item['Descriptive Properties'] = self.get_properties(item, response, 'Descriptive','Descriptive Properties')
            self.get_properties(item, response, 'Descriptive', 'Descriptive Properties')
            # item['Chemical Resistance Properties'] = self.get_properties(item, response, 'Chemical Resistance','Chemical Resistance Properties')
            self.get_properties(item, response, 'Chemical Resistance', 'Chemical Resistance Properties')
            item['URL'] = url

            if not title:
                self.write_logs(f'Wrong Response get from URL: {url}')
                return

            self.current_sub_cat_scraped += 1
            self.all_groups_scraped_records += 1
            print(f'all_groups_scraped_records: {self.all_groups_scraped_records}')
            self.read_write_xlsx_file(record=item, key='record')
        except Exception as e:
            self.write_logs(f'Error in item write URL:{url} Error:{e}')

    def normalize_url(self, url):
        """Remove 'ckck' query parameter from URLs to standardize them."""
        parsed_url = urlparse(url)
        query_params = parse_qs(parsed_url.query)

        # Remove 'ckck' parameter if it exists
        query_params.pop('ckck', None)

        # Rebuild the URL without 'ckck'
        new_query = urlencode(query_params, doseq=True)
        normalized_url = f"{parsed_url.scheme}://{parsed_url.netloc}{parsed_url.path}"
        if new_query:
            normalized_url += f"?{new_query}"

        return normalized_url

    def get_properties(self, item, response, key, value):
        try:
            # Extract the headers (keys)
            keys = [
                       key.strip() for key in response.css(f'tr:contains("{value} Properties") th ::text').getall()[1:]
                   ] or response.css(f'tr:contains("{value}") th::text').getall()[1:]
            if value == 'Descriptive Properties':
                keys = response.css(f'tr:contains("Descriptive Properties") th ::text').getall()
                if keys:
                    keys = ['Descriptive Properties']
            if not keys:
                return 'N/A'  # Return early if no keys are found

            # Extract all rows following the specified property section
            rows = response.css(f'tr:contains("{value} Properties") ~ tr')
            if not rows:
                rows = response.css(f'tr:contains("{value}") ~ tr')

            # Initialize a variable to track the last valid main key
            last_main_key = None

            for row in rows:
                cells = row.css('td')
                if 'colspan="4"' in cells.get(''):
                    break

                if cells:
                    if 'Descriptive Properties' in keys:
                        main_key = row.css('td::text').get('')
                        if not main_key:
                            continue
                    else:
                        # Extract property name and corresponding values
                        main_key = cells[0].css('::text').get('').strip()  # First column is the property name

                        # Use the last valid main key if the current one is empty
                        if not main_key:
                            main_key = last_main_key
                        else:
                            last_main_key = main_key  # Update the last valid main key

                    # If there's still no main key, skip this row
                    if not main_key:
                        print('Skipping row with no main key')
                        continue

                    property_key = f'{key}-{main_key}'
                    if property_key not in self.fields:
                        self.fields.append(property_key)

                    remaining_values = list(
                        filter(None, (''.join(cell.css('::text').getall()).strip() for cell in cells[1:])))

                    # Map the extracted values to the keys
                    if 'Descriptive Properties' in keys:
                        item[property_key] = {main_key: remaining_values[0] if remaining_values else ''}
                    else:
                        item[property_key] = {
                            main_key: {
                                keys[i]: remaining_values[i] if i < len(remaining_values) else None
                                for i in range(len(keys))
                            }
                        }

        except Exception as e:
            # Catch any exception and return 'N/A'
            print(f"Error in get_properties: {e}")  # Optionally log the error for debugging
            return 'N/A'

    def get_physical_properties(self, item, response, value):
        try:
            # Extract the headers (keys)
            keys = [
                       key.strip() for key in response.css(f'tr:contains("{value} Properties") th ::text').getall()[1:]
                   ] or response.css(f'tr:contains("{value}") th::text').getall()[1:]

            if not keys:
                return 'N/A'  # Return early if no keys are found

            # Extract all rows following the specified property section
            rows = response.css(f'tr:contains("{value} Properties") ~ tr')
            if not rows:
                rows = response.css(f'tr:contains("{value}") ~ tr')

            # Initialize a variable to track the last valid main key
            last_main_key = None

            # Process each row and extract the data
            property_dicts = []
            for row in rows:
                cells = row.css('td')
                if 'colspan="4"' in cells.get(''):
                    break

                if cells:
                    # Extract property name and corresponding values
                    main_key = cells[0].css('::text').get('').strip()  # First column is the property name

                    # Use the last valid main key if the current one is empty
                    if not main_key:
                        main_key = last_main_key
                    else:
                        last_main_key = main_key  # Update the last valid main key

                    # If there's still no main key, skip this row
                    if not main_key:
                        print('Skipping row with no main key')
                        continue

                    remaining_values = [
                        ''.join(cell.css('::text').getall()).strip() for cell in cells[1:]
                    ]

                    key = f'Physical-{main_key}'
                    if key not in self.fields:
                        self.fields.append(key)

                    item[key] = {
                        keys[i]: remaining_values[i] if i < len(remaining_values) else None
                        for i in range(len(keys))
                    }

        except Exception as e:
            # Catch any exception and return 'N/A'
            print(f"Error in get_properties: {e}")  # Optionally log the error for debugging
            return 'N/A'

    def get_vendor(self, response):
        vendors_text = response.css('tr:contains("Vendors") td ::text').getall()
        if any("No vendors are listed for this" in text for text in vendors_text):
            return 'N/A'
        else:
            text = ' '.join(vendors_text).split('Click here  to view all')[0]

            return text

    def get_sub_cat_form_data(self, viewstate, view_genrator, parent_category, sub_category_dict, key, page_no=None):
        sub_cat_id = sub_category_dict.get('MatGroupID', 0)
        name = sub_category_dict.get('GroupName', '')
        count = sub_category_dict.get('MatCount', '')
        sub_category_text = f'{name} ({count} matls)'

        if key == 'first_page':
            data = {
                '__EVENTTARGET': '',
                '__EVENTARGUMENT': '',
                '__LASTFOCUS': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_ExpandState': 'ccccccnc',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_SelectedNode': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_PopulateLog': '',
                '__VIEWSTATE': viewstate,
                '__VIEWSTATEGENERATOR': view_genrator,
                'ctl00$txtQuickText': '',
                'ctl00$ContentMain$ucPopupMessage1$hndPopupControl': '',
                'ctl00$ContentMain$UcMatGroupFinder1$txtSearchText': parent_category,
                'ctl00$ContentMain$UcMatGroupFinder1$selectCategoryList': str(sub_cat_id),
                'ctl00$ContentMain$UcMatGroupFinder1$TextBoxWatermarkExtender2_ClientState': '',
                'ctl00$ContentMain$txtMatGroupID': str(sub_cat_id),
                'ctl00$ContentMain$txtMatGroupText': sub_category_text,
                'ctl00$ContentMain$btnSubmit.x': '25',
                'ctl00$ContentMain$btnSubmit.y': '12',
            }
            return data
        elif key == 'two_hundred':
            data = {
                '__EVENTTARGET': 'ctl00$ContentMain$UcSearchResults1$drpPageSize1',
                '__EVENTARGUMENT': '',
                '__LASTFOCUS': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_ExpandState': 'ccccccnc',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_SelectedNode': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_PopulateLog': '',
                '__VIEWSTATE': viewstate,
                '__VIEWSTATEGENERATOR': '640EFF8C',
                'ctl00$txtQuickText': '',
                'ctl00$ContentMain$ucPopupMessage1$hndPopupControl': '',
                'ctl00$ContentMain$UcMatGroupFinder1$txtSearchText': parent_category,
                'ctl00$ContentMain$UcMatGroupFinder1$TextBoxWatermarkExtender2_ClientState': '',
                'ctl00$ContentMain$txtMatGroupID': str(sub_cat_id),
                'ctl00$ContentMain$txtMatGroupText': sub_category_text,
                # 'ctl00$ContentMain$UcSearchResults1$drpPageSelect1': '1',
                'ctl00$ContentMain$UcSearchResults1$drpPageSelect1': str(page_no),
                'ctl00$ContentMain$UcSearchResults1$drpPageSize1': '200',
                'ctl00$ContentMain$UcSearchResults1$drpFolderList': '0',
                'ctl00$ContentMain$UcSearchResults1$txtFolderMatCount': '0/0',
                # 'ctl00$ContentMain$UcSearchResults1$drpPageSelect2': '1',
                'ctl00$ContentMain$UcSearchResults1$drpPageSelect2': str(page_no),
                'ctl00$ContentMain$UcSearchResults1$drpPageSize2': '50',
            }
            return data
        # Define and incorporate next_page data if `next_page` is True
        elif key=='pagination':  #pagination
            data = {
                # '__EVENTTARGET': 'ctl00$ContentMain$UcSearchResults1$lnkNextPage',
                '__EVENTTARGET': 'ctl00$ContentMain$UcSearchResults1$drpPageSize1',
                '__EVENTARGUMENT': '',
                '__LASTFOCUS': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_ExpandState': 'ccccccnc',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_SelectedNode': '',
                'ctl00_ContentMain_ucMatGroupTree_msTreeView_PopulateLog': '',
                '__VIEWSTATE': viewstate,
                '__VIEWSTATEGENERATOR': '640EFF8C',
                'ctl00$txtQuickText': '',
                'ctl00$ContentMain$ucPopupMessage1$hndPopupControl': '',
                'ctl00$ContentMain$UcMatGroupFinder1$txtSearchText': str(parent_category),
                'ctl00$ContentMain$UcMatGroupFinder1$TextBoxWatermarkExtender2_ClientState': '',
                'ctl00$ContentMain$txtMatGroupID': str(sub_cat_id),
                'ctl00$ContentMain$txtMatGroupText': sub_category_text,
                'ctl00$ContentMain$UcSearchResults1$drpPageSelect1': str(page_no),
                'ctl00$ContentMain$UcSearchResults1$drpPageSize1': '200',
                'ctl00$ContentMain$UcSearchResults1$drpFolderList': '0',
                'ctl00$ContentMain$UcSearchResults1$txtFolderMatCount': '0/0',
                'ctl00$ContentMain$UcSearchResults1$drpPageSelect2': str(page_no),
                'ctl00$ContentMain$UcSearchResults1$drpPageSize2': '200',
            }
            return data

    def read_json_files(self, filepath):
        files = glob.glob(filepath)

        data = []
        for file in files:
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    file_data = json.load(f)
                    data.append(file_data)

            except json.JSONDecodeError as e:
                print(f"Error decoding JSON in file {file}: {e}")
            except Exception as e:
                print(f"Unexpected error while processing file {file}: {e}")

        return data

    def get_subcategories(self, categories):
        headers = {
            'accept': '*/*',
            'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
            'content-type': 'application/json; charset=UTF-8',
            'origin': 'https://www.matweb.com',
            'priority': 'u=1, i',
            'referer': 'https://www.matweb.com/search/MaterialGroupSearch.aspx',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
        }
        sub_categories = []

        for category in categories:
            json_data = {'prefixText': category, }

            resp = requests.post('https://www.matweb.com/WebServices/MatGroups.asmx/FindMatGroups', headers=headers,
                                 json=json_data, )

            list_dict = resp.json()
            for row in list_dict:
                row['parent_category'] = category  # Add parent_category field
                sub_categories.append(row)

        return sub_categories

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def read_skipped_urls(self):
        """Read skipped URLs from the file, if available."""
        # Ensure the file exists
        if not os.path.exists(self.skipped_urls_filepath):
            with open(self.skipped_urls_filepath, 'w', encoding='utf-8') as file:
                return []  # Create the file if it doesn't exist

        try:
            with open(self.skipped_urls_filepath, 'r', encoding='utf-8') as file:
                return [line.strip() for line in file if line.strip()]
        except (IndexError, FileNotFoundError):
            self.write_logs("Skipped URLs file not found.")
        except Exception as e:
            self.write_logs(f"Error reading skipped URLs: {e}")
        return []

    def write_skipped_urls(self, log_msg):
        # Ensure the file exists
        if not os.path.exists(self.skipped_urls_filepath):
            with open(self.skipped_urls_filepath, 'w', encoding='utf-8') as logs_file:
                pass  # Create the file if it doesn't exist

        # Read and write operations ensuring full-line match
        with open(self.skipped_urls_filepath, 'r+', encoding='utf-8') as logs_file:
            # Check if the exact log_msg exists in the file
            if any(line.strip() == log_msg for line in logs_file):
                print(f'{log_msg} already in Skipped URLs. Skipped...')
                return
            # Append the log_msg to the file
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def  read_write_xlsx_file(self, record=None, key=None):
        json_file = f'logs/{self.name} Materials Details.json'
        try:
            if key == 'previous_records':
                file_exists = os.path.exists(self.output_file_path)
                if not file_exists:
                    return {}

                # Load the workbook (assume the first match)
                workbook = load_workbook(self.output_file_path)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]

                sheet_headers = list(sorted(set(headers)))
                for header in sheet_headers:  #update the headers from previous scraped Excel file
                    if header not in self.fields:
                        self.fields.append(header)

                data = {}
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    dict_row = dict(zip(headers, row))
                    url = dict_row.get('URL', '')
                    data[url] = dict_row
                    self.ready_write_records.append(dict_row)
                    self.ready_write_records_counter += 1
                return data

            elif key == 'record' or key == 'close_spider':
                if record:
                    self.ready_write_records.append(record)
                self.ready_write_records_counter += 1

                # Only write to the file when we have 1000 records
                if len(self.ready_write_records) >= 1000 or key == 'close_spider':
                    try:
                        # first create json file temporary to enhance the script performance
                        with open(json_file, 'r', encoding='utf-8') as f:
                            existing_data = json.load(f)
                    except FileNotFoundError:
                        existing_data = []

                    existing_data.extend(self.ready_write_records)

                    with open(json_file, 'w', encoding='utf-8') as f:
                        json.dump(existing_data, f, indent=4, ensure_ascii=False)

                    self.ready_write_records = []
                    if key == 'close_spider':
                        # first read all records from the JSON File json_file
                        with open(json_file, 'r', encoding='utf-8') as f:
                            all_records = json.load(f)

                        file_exists = os.path.exists(self.output_file_path)
                        # If the file exists, delete it
                        if file_exists:
                            os.remove(self.output_file_path)

                        # Create a new workbook and add headers
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Materials"
                        ws.append(self.fields)
                        for record in all_records:
                            processed_record = []
                            for field in self.fields:
                                value = record.get(field, 'N/A')

                                # If the value is a list of dictionaries, convert it to a human-readable string
                                if isinstance(value, list):
                                    value = ",\n ".join([", ".join(f"{k}: {v}" for k, v in item.items()) if isinstance(
                                        item, dict) else str(item) for item in value])

                                    # Convert dictionaries to a readable format
                                elif isinstance(value, dict):
                                    value = ", ".join(f"{k}: {v}" for k, v in value.items() if v)

                                # Handle empty or invalid values
                                if value in (None, '', ' '):
                                    value = 'N/A'

                                processed_record.append(value)
                            ws.append(processed_record)

                        # Save the updated workbook to the file
                        wb.save(self.output_file_path)
                        os.remove(json_file)

        except Exception as e:
            self.write_logs(f"An error occurred: {e}")
            return {}

    def close(Spider, reason):
        Spider.read_write_xlsx_file(record=None, key='close_spider')

        # Log overall scraping statistics
        Spider.write_logs(f"\n--- Scraping Summary ---")
        Spider.write_logs(f"Total Products Available on Website: {Spider.all_groups_found_records}")
        Spider.write_logs(f"Total Products Successfully Scraped: {Spider.all_groups_scraped_records}")

        # Log script execution times
        Spider.write_logs(f"\n--- Script Execution Times ---")
        Spider.write_logs(f"Script Start Time: {Spider.script_starting_datetime}")
        Spider.write_logs(f"Script End Time: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        Spider.write_logs(f"Reason for Closure: {reason}")

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(MatwebSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        if self.current_scraped_sub_cat:
            self.write_logs(
                f"\n\n{self.current_scraped_sub_cat} Founded Records:{self.current_sub_cat_found} and Scraped Successfully:{self.current_sub_cat_scraped}\n\n")
            self.write_skipped_urls(f'{self.current_scraped_sub_cat}')
            self.current_sub_cat_found = 0
            self.current_sub_cat_scraped = 0
            self.current_scraped_sub_cat = ''

        if self.all_sub_categories:
            # Log progress
            self.write_logs(
                f"\n\n{len(self.all_sub_categories)}/{self.all_material_groups} Material Group Subcategories left to Scrape\n\n")

            sub_category_dict = self.all_sub_categories.pop()
            sub_category = sub_category_dict.get('GroupName', '').strip()
            self.current_scraped_sub_cat = sub_category
            parent_category = sub_category_dict.get('parent_category', '').title()

            info = f'Parent Category: {parent_category} Sub Category: {sub_category}'
            self.write_logs(f'{info} Starting for Scraped')
            if sub_category in self.skipped_urls:  # if Material Group Already Scraped
                self.write_logs(f'{info} already scraped last time.')
                self.current_scraped_sub_cat = ''
                self.spider_idle()
                return

            self.crawler.engine.crawl(Request(url='https://books.toscrape.com', callback=self.parse_home,
                                              dont_filter=True, meta={'sub_category': sub_category_dict}))
