import glob
import json
import os
from datetime import datetime

import requests
from openpyxl.reader.excel import load_workbook

from urllib import parse
from scrapy import Request, Spider, Selector
from scrapy.exceptions import CloseSpider


class TaxSpider(Spider):
    name = "tax"
    allowed_domains = ["www.taxappeals.ie"]
    start_urls = ["https://www.taxappeals.ie/en/determinations"]
    current_dt = datetime.now().strftime("%d%m%Y%H%M")

    custom_settings = {
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        # 'CONCURRENT_REQUESTS': 3,
    }

    def __init__(self):
        super().__init__()
        # Logs & tax_appeals folders
        os.makedirs('logs', exist_ok=True)
        os.makedirs('tax_appeals', exist_ok=True)

        self.logs_filepath = f'logs/{self.name}_logs {self.current_dt}.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'[INIT] Script started at {self.script_starting_datetime}')

        self.read_determination_database= self.read_xlsx_input_file()
        self.previous_scraped_records = self.previous_records()
        self.product_count = 0

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    }

    def start_requests(self):
        self.write_logs(f"\nTotal Records found in the Excel file determination_database are :{len(self.read_determination_database)}\n")
        for record in self.read_determination_database:

            try:
                publication_ref = record.get('Publication Ref', '').strip()

                #previous scraped records checks
                if publication_ref in self.previous_scraped_records:
                    print(f'Publication Ref {publication_ref} Already Scraped so Skipped...')
                    self.product_count += 1
                    print('Items Scrapped:', self.product_count)
                    continue

                url = record.get('Link','').strip()

                if 'xlsx' in url or 'VIEW' in url:
                    self.product_count += 1
                    print('Items Scrapped:', self.product_count)
                    self.write_logs(f"Publication Ref: {publication_ref}  not has valid URL")
                    continue

                if '/Microsoft/' in url:
                    print('url', url)
                    url = ''.join(url.split('taxappeals.ie/')[1:])
                    url = f'https://www.taxappeals.ie/{url}'

                url = url[:-1] if url.endswith('-') else url
                yield Request(url, callback=self.parse_detail_page, dont_filter=True,
                              headers=self.headers, meta={'record':record})

            except Exception as e:
                a=1

        # homepage request
        yield Request(url=self.start_urls[0], callback=self.parse, headers=self.headers)

    def parse(self, response, **kwargs):
        determinations_urls = response.css('.blogSummary a::attr(href)').getall() or []
        for url in determinations_urls:
            try:
                pub_name = ''.join(''.join(url.split('/')[-1:]).split('-')[:1]).upper()
                if pub_name in self.previous_scraped_records:
                    print(f'Publication Ref {pub_name} Already Scraped so Skipped...')
                    self.product_count += 1
                    print('Items Scrapped:', self.product_count)
                    continue
            except:
                a=1

            url = f'https://www.taxappeals.ie{url}'
            yield Request(url, callback=self.parse_detail_page, headers=self.headers)

        if not response.meta.get('pagination', ''):
            for page_no in range(2, 21):
                url = f'https://www.taxappeals.ie/en/determinations/{page_no}'
                yield Request(url, callback=self.parse, headers=self.headers, meta={'pagination':True})

    def parse_detail_page(self, response):
        record = response.meta.get('record', '')
        if not record:  # if record empty then it mean response from 2024 year web base request
            record = self.get_record_dict(response)

        if 'encountered an unexpected URL' in response.text:
            url = response.url.replace('/-', '/')
            req = requests.get(url)
            if req.status_code==200:
                response = Selector(text=req.text)
            else:
                self.write_logs(f"Records Not Exist at detail page URL:{response.url}")
                return

        try:
            pdf_urls = response.css('#contentDiv a::attr(href)').getall() or []

            for pdf_url in pdf_urls:
                pdf_filename = ''
                if pdf_url:
                    pdf_url = f'https://www.taxappeals.ie{pdf_url}'

                    # Download the PDF and get the file name
                    pdf_filename = self.download_pdf_content(pdf_url)

                pdf_filename = pdf_filename if pdf_filename else f"{record.get('Publication Ref', '')}.PDF"
                save_json = self.write_json(record, pdf_filename)

                self.product_count += 1
                print('Items Scrapped: ', self.product_count)

        except Exception as e:
            self.write_logs(f"Error processing detail page: {str(e)}")

    def read_xlsx_input_file(self):
        try:
            filename = glob.glob('input/*.xlsx')[0]

            # Check if a file is found
            if not filename:
                self.write_logs("Input file not found in the /input directory.")
                raise CloseSpider("Input file not found. Closing spider.")

            # Load the workbook (assume the first match)
            workbook = load_workbook(filename)

            # Select the active worksheet
            sheet = workbook.active

            # Extract the header (first row)
            headers = [cell.value for cell in sheet[1]]

            data = []
            for row in sheet.iter_rows(min_row=2):  # Iterate over row objects, not just values
                row_data = {}
                for header, cell in zip(headers, row):
                    if cell.hyperlink:  # Check if the cell has a hyperlink
                        row_data[header] = cell.hyperlink.target  # Extract the actual URL
                    else:
                        row_data[header] = cell.value  # Use the cell's value if no hyperlink exists

                # Check if the row contains a year (e.g., 2016) before appending
                if any(isinstance(value, int) and 1900 <= value <= 2100 for value in row_data.values()):
                    data.append(row_data)

            return data

        except Exception as e:
            self.write_logs(f"An error occurred: {e}")
            raise CloseSpider("An error occurred while reading the input file. Closing spider.")

    def download_pdf_content(self, url):
        try:
            folder_path = 'tax_appeals'
            file_name = url.split("/")[-1]
            file_name = parse.unquote(file_name)
            file_name = file_name.replace("Determination", "").replace('-', '').strip()
            file_name = file_name.replace("(1)", "").strip().upper()

            # Define the full file path
            file_path = os.path.join(folder_path, file_name)

            # Check if the file already exists
            if os.path.exists(file_path):
                print(f"File '{file_name}' already exists. Skipping download.")
                return file_name

            # Send a GET request to the URL
            response = requests.get(url, stream=True)

            # Check if the request was successful
            if response.status_code == 200 and 'application/pdf' in response.headers.get('Content-Type', ''):
                # Save the PDF file locally in the specified folder
                file_path = os.path.join(folder_path, file_name)
                with open(file_path, 'wb') as pdf_file:
                    for chunk in response.iter_content(chunk_size=1024):
                        pdf_file.write(chunk)

                print(f"PDF successfully downloaded and saved as: {file_name}")
                return file_name
            else:
                self.write_logs(f"Failed to download PDF. HTTP Status: {response.status_code}, Content-Type: {response.headers.get('Content-Type')}")
                return ''
        except Exception as e:
            self.write_logs(f"An error occurred while downloading the PDF: {e}")
            return ''

    def write_json(self, record, pdf_file):
        try:
            # Define the folder path
            folder_path = 'tax_appeals'

            # Create the folder if it doesn't exist
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            # Construct the JSON file name
            json_file_name = f"{pdf_file}.metadata.json"
            file_path = os.path.join(folder_path, json_file_name)

            if "Link" in record:
                record.pop("Link")

            # Write the record to the JSON file
            with open(file_path, 'w') as json_file:
                json.dump(record, json_file, indent=4)

            print(f"JSON metadata successfully written to: {file_path}")
            return file_path

        except Exception as e:
            print(f"An error occurred while writing the JSON metadata: {e}")

    def previous_records(self):
        try:
            pdf_files = glob.glob('tax_appeals/*.pdf')
            json_files = glob.glob('tax_appeals/*.json')

            # Create a list to store the matching PDF and JSON file pairs
            matching_files = []

            # Check for matching files where the PDF name is the initial part of the JSON file name
            for pdf_file in pdf_files:
                pdf_file_name = os.path.basename(pdf_file).replace(".PDF", "")
                for json_file in json_files:
                    json_file_name = os.path.basename(json_file).replace(".json", "").split(".")[0]
                    if pdf_file_name == json_file_name:
                        matching_files.append((pdf_file, json_file))

            # Extract the PDF names (without the ".pdf" extension) for the matching files
            pdf_names = [os.path.basename(pdf_file).replace(".pdf", "") for pdf_file, _ in matching_files]

            return [name.replace('.PDF', '').strip() for name in pdf_names]

        except Exception as e:
            print(f"An error occurred while processing the files: {e}")
            return []

    def get_record_dict(self, response):
        tax_heads_mapping = {
            'VAT':'VAT',
            'Capital Gains Tax': 'CGT',
            'Capital Gains Tax & Income Tax': 'CGT & IT',
            'CGT': 'CGT',
            'Corporation Tax': 'CT',
            'Customs and Excise': 'CE',
            'Customs & Excise' : 'CE',
            'Corporation Tax & Income Tax': 'CT & IT',
            'CT, DWT & PREM': 'CT & DWT & PREM',
            'Deposit Interest Retention Tax': 'DIRT',
            'Employment and Investment Incentive': 'EII',
            'CGT, Income Tax': 'CGT & IT',
            'Income Tax': 'IT',
            'Income Tax & VAT' : 'IT & VAT',
            'Income Tax & RCT' : 'IT & RCT',
            'Income Tax & USC': 'IT & USC',
            'Income Tax & Capital Gains Tax': 'IT & CGT',
            'Income Tax & CAT': 'IT & CAT',
            'Income Tax - PSWT': 'IT & PSWT',
            'Income Tax and USC': 'IT & USC',
            'Income Tax & CGT': 'IT & CGT',
            'Income Tax (Rental or Trading Income)': 'IT (R or T)',
            'Income Tax (Help to Buy)': 'IT (Help to Buy)',
            'Income Tax & LPT': 'IT & LPT',
            'Income Tax – Undeclared Income – CAB': 'IT - UI - CAB',
            'CGT & Income Tax': 'CGT & IT',
            'Local Property Tax': 'LPT',
            'LPT': 'LPT',
            'Pay As You Earn': 'PAYE',
            'Relevant Contracts Tax': 'RCT',
            'Special Assignee Relief Programme': 'SARP',
            'Stamp Duty': 'Stamp Duty',
            'Temporary Wage Subsidy Scheme': 'TWSS',
            'Universal Social Charge': 'USC',
            'Value Added Tax': 'VAT',
            'Vacant Homes Tax': 'VHT',
            'VAT & Customs and Excise': 'VAT & CE',
            'Vehicle Registration Tax': 'VRT',
            'Vehicle Registration Tax & Value Added Tax': 'VRT & VAT',
            'Value Added Tax & Vehicle Registration Tax': 'VAT & VRT',
            'Employment Investment Incentive': 'EII',
            'PAYE, PRSI & USC': 'PAYE, PRSI & USC',
            'Capital Acquisitions Tax': 'CAT',
            'Help to Buy Scheme': 'HB Scheme',
            'Artists Exemption': 'AE',
            "Artists' Exemption" : 'AE',
            'EWSS': 'EWSS',
            'Artist Exemption': 'AE',
            'VRT': 'VRT',
            'Covid Relief - EWSS': 'CR & EWSS',
            'CRSS': 'CRSS',
            'PREM': 'PREM',
            'CAT': 'CAT',
            'RCT': 'RCT',
            'SPCCC': 'SPCCC'
        }

        try:
            title = response.css('#content h1::text').get('').strip()
            publication_ref = title.split('-')[0].strip()
            year = ''

            if '2024' in publication_ref:
                year = '2024'
            else:
                try:
                    year = publication_ref.split('–')[0][-5:].strip()
                except:
                    year= publication_ref.split('-')[0][-5:].strip()

            raw_tax_head = title.split('-')[1].strip()
            tax_head = tax_heads_mapping.get(raw_tax_head, 'Unknown')  # Default to 'Unknown' if not found

            if tax_head=='Unknown':
                tax_head = raw_tax_head
                self.write_logs(f'Tax Head Mismatch URL:{response.url}')

            dict_record =  {
                'Year': year,
                'Publication Ref': publication_ref,
                'Tax head': tax_head,
                "Topic Summary": '',
                "Tax Legislation": ''
            }

            return dict_record
        except:
            return {}

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def close(Spider, reason):
        Spider.write_logs(f'\n\nTotal Items Scraped:{Spider.product_count}')
        Spider.write_logs(f'Spider Started from :{Spider.script_starting_datetime}')
        Spider.write_logs(f'Spider Stopped at :{datetime.now().strftime("%d-%m-%Y %H:%M:%S")}')
